import customtkinter as ctk
import database.db_manager as db
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import os
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

class CreditLedgerView(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")
        
        self.active_customer = None # Focused Customer Node Tuple Cache
        self.current_page = 1
        self.page_size = 10

        self.pb_current_page = 1
        self.pb_page_size = 25
        
        # --- SUB-TAB CONTROL HEADER STRIP ---
        self.tabs_nav_frame = ctk.CTkFrame(self, height=45, fg_color="#16161a", border_color="#222227", border_width=1)
        self.tabs_nav_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.btn_tab_directory = ctk.CTkButton(self.tabs_nav_frame, text="💳 Wallet Directory", width=160, height=35, fg_color="#1f1f24", font=ctk.CTkFont(weight="bold"), command=self.switch_to_directory_tab)
        self.btn_tab_directory.pack(side="left", padx=10, pady=5)
        
        self.btn_tab_passbook = ctk.CTkButton(self.tabs_nav_frame, text="📖 Passbook Timeline", width=160, height=35, fg_color="transparent", text_color="gray", font=ctk.CTkFont(weight="bold"), command=self.switch_to_passbook_tab)
        self.btn_tab_passbook.pack(side="left", padx=5, pady=5)
        
        self.btn_tab_cashflow = ctk.CTkButton(self.tabs_nav_frame, text="📊 Agency Cash Flow", width=160, height=35, fg_color="transparent", text_color="gray", font=ctk.CTkFont(weight="bold"), command=self.switch_to_cashflow_tab)
        self.btn_tab_cashflow.pack(side="left", padx=5, pady=5)

        # --- CENTRAL WORKSPACE VIEWS HOLDER ---
        self.view_container = ctk.CTkFrame(self, fg_color="transparent")
        self.view_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Initialize sub-views layout blocks
        self.init_directory_tab_ui()
        self.init_passbook_tab_ui()
        self.init_cashflow_tab_ui()
        
        # Default Tab Launch Context
        self.switch_to_directory_tab()

    # =================================================================
    # 🎛️ TAB ROUTING ROUTINES
    # =================================================================
    def switch_to_directory_tab(self):
        self.passbook_frame.pack_forget()
        self.cashflow_frame.pack_forget()
        self.directory_frame.pack(fill="both", expand=True)
        
        self.btn_tab_directory.configure(fg_color="#1f1f24", text_color="#ffffff")
        self.btn_tab_passbook.configure(fg_color="transparent", text_color="gray")
        self.btn_tab_cashflow.configure(fg_color="transparent", text_color="gray")
        self.fetch_directory_data()

    def switch_to_passbook_tab(self):
        if not self.active_customer:
            messagebox.showwarning("Selection Missing", "Please focus on a customer from the Wallet Directory first to pull statement logs.")
            self.switch_to_directory_tab()
            return
        self.directory_frame.pack_forget()
        self.cashflow_frame.pack_forget()
        self.passbook_frame.pack(fill="both", expand=True)
        
        self.btn_tab_directory.configure(fg_color="transparent", text_color="gray")
        self.btn_tab_passbook.configure(fg_color="#1f1f24", text_color="#ffffff")
        self.btn_tab_cashflow.configure(fg_color="transparent", text_color="gray")
        self.pb_current_page = 1
        self.render_passbook_timeline()

    def switch_to_cashflow_tab(self):
        self.directory_frame.pack_forget()
        self.passbook_frame.pack_forget()
        self.cashflow_frame.pack(fill="both", expand=True)
        
        self.btn_tab_directory.configure(fg_color="transparent", text_color="gray")
        self.btn_tab_passbook.configure(fg_color="transparent", text_color="gray")
        self.btn_tab_cashflow.configure(fg_color="#1f1f24", text_color="#ffffff")
        self.render_cashflow_metrics()

    # =================================================================
    # 💳 SUB-TAB 1: WALLET MASTER DIRECTORY VIEW BUILDER
    # =================================================================
    def init_directory_tab_ui(self):
        self.directory_frame = ctk.CTkFrame(self.view_container, fg_color="transparent")
        self.directory_frame.grid_columnconfigure(0, weight=4)
        self.directory_frame.grid_columnconfigure(1, weight=3)
        self.directory_frame.grid_rowconfigure(0, weight=1)
        
        # Left Panel (Directory List View Grid)
        left_pane = ctk.CTkFrame(self.directory_frame, fg_color="transparent")
        left_pane.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        search_box = ctk.CTkFrame(left_pane, fg_color="#16161a", border_color="#222227", border_width=1, corner_radius=6)
        search_box.pack(fill="x", pady=(0, 10))
        
        self.search_entry = ctk.CTkEntry(search_box, placeholder_text="🔍 Search by Khata No. (KH-0004), Name, or Phone...", height=38, fg_color="#121214", border_color="#222227")
        self.search_entry.pack(side="left", fill="x", expand=True, padx=15, pady=10)
        self.search_entry.bind("<KeyRelease>", lambda e: self.on_search_changed())
        
        ctk.CTkButton(search_box, text="📊 Export All Khata", height=38, width=145,
                      fg_color="#1c3a27", text_color="#4aff4a",
                      border_color="#2d5a3b", border_width=1,
                      font=ctk.CTkFont(weight="bold"),
                      command=self.export_all_khata_to_xlsx).pack(side="right", padx=6, pady=10)

        ctk.CTkButton(search_box, text="📄 Export PDF", height=38, width=125,
                      fg_color="#3a1c1c", text_color="#ff9f9f",
                      border_color="#5a2d2d", border_width=1,
                      font=ctk.CTkFont(weight="bold"),
                      command=self.export_all_khata_to_pdf).pack(side="right", padx=6, pady=10)

        ctk.CTkButton(search_box, text="📥 Import Sheet", height=38, width=130,
                      fg_color="#1f293d", text_color="#8ed1fc",
                      border_color="#2b3a4a", border_width=1,
                      font=ctk.CTkFont(weight="bold"),
                      command=self.import_khata_from_xlsx).pack(side="right", padx=6, pady=10)
        
        ctk.CTkButton(search_box, text="+ Open Wallet", height=38, width=120, fg_color="#1f293d", font=ctk.CTkFont(weight="bold"), command=self.open_new_wallet_modal).pack(side="right", padx=(0,4), pady=10)
        
        self.dir_table_scroll = ctk.CTkScrollableFrame(left_pane, fg_color="#121214", border_color="#222227", border_width=1, corner_radius=6)
        self.dir_table_scroll.pack(fill="both", expand=True)

        # Pagination Bar
        pagination_bar = ctk.CTkFrame(left_pane, fg_color="#16161a", border_color="#222227", border_width=1, corner_radius=6, height=44)
        pagination_bar.pack(fill="x", pady=(10, 0))

        self.btn_prev_page = ctk.CTkButton(pagination_bar, text="⬅ Previous", width=100, height=32, fg_color="#1f1f24", font=ctk.CTkFont(size=12), command=self.go_prev_page)
        self.btn_prev_page.pack(side="left", padx=10, pady=6)

        self.lbl_page_info = ctk.CTkLabel(pagination_bar, text="Page 1 of 1", font=ctk.CTkFont(size=12, weight="bold"), text_color="#a0a0a9")
        self.lbl_page_info.pack(side="left", expand=True)

        self.btn_next_page = ctk.CTkButton(pagination_bar, text="Next ➡", width=100, height=32, fg_color="#1f1f24", font=ctk.CTkFont(size=12), command=self.go_next_page)
        self.btn_next_page.pack(side="right", padx=10, pady=6)

        # Right Panel (Operations Management Dock terminal Console)
        right_pane = ctk.CTkFrame(self.directory_frame, fg_color="#16161a", border_color="#222227", border_width=1, corner_radius=6)
        right_pane.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        ctk.CTkLabel(right_pane, text="Wallet Console Desk", font=ctk.CTkFont(size=14, weight="bold"), text_color="#4a90e2").pack(pady=15, anchor="w", padx=20)
        self.lbl_focused_name = ctk.CTkLabel(right_pane, text="No Profile Focused", font=ctk.CTkFont(size=18, weight="bold"), text_color="#ffffff")
        self.lbl_focused_name.pack(anchor="w", padx=20, pady=(0, 15))
        
        # Balance Tile Card Block
        ctk.CTkLabel(right_pane, text="Current Advance Deposit Standing:", font=ctk.CTkFont(size=11), text_color="gray").pack(anchor="w", padx=20)
        self.lbl_wallet_balance_tile = ctk.CTkLabel(right_pane, text="Rs. 0.00", font=ctk.CTkFont(size=24, weight="bold"), text_color="gray")
        self.lbl_wallet_balance_tile.pack(anchor="w", padx=20, pady=(0, 20))
        
        # Transaction Inputs Section Area Box
        input_card = ctk.CTkFrame(right_pane, fg_color="#121214", border_color="#222227", border_width=1, corner_radius=4)
        input_card.pack(fill="x", padx=20, pady=10)
        
        self.inp_op_amount = ctk.CTkEntry(input_card, placeholder_text="Enter Ledger Cash Figure Amount (Rs.) *", height=38, fg_color="#16161a", border_color="#222227")
        self.inp_op_amount.pack(fill="x", padx=15, pady=(15, 8))
        
        self.inp_op_desc = ctk.CTkEntry(input_card, placeholder_text="Narration Context Details (e.g., Token Advance, Refund)", height=38, fg_color="#16161a", border_color="#222227")
        self.inp_op_desc.pack(fill="x", padx=15, pady=8)
        
        # Buttons Setup Actions Row Split
        self.btn_deposit = ctk.CTkButton(input_card, text="📥 Deposit Advance", fg_color="#1c3a27", text_color="#4aff4a", height=38, font=ctk.CTkFont(weight="bold"), state="disabled", command=lambda: self.execute_operation("ADVANCE_DEPOSIT"))
        self.btn_deposit.pack(fill="x", padx=15, pady=4)
        
        self.btn_withdraw = ctk.CTkButton(input_card, text="💸 Withdraw Cash", fg_color="#3a1c1c", text_color="#ff8080", height=38, font=ctk.CTkFont(weight="bold"), state="disabled", command=lambda: self.execute_operation("CASH_WITHDRAWAL"))
        self.btn_withdraw.pack(fill="x", padx=15, pady=4)
        
        self.btn_purchase_adj = ctk.CTkButton(input_card, text="🛒 Adjust Purchase Debit", fg_color="#1f293d", text_color="#8ed1fc", height=38, font=ctk.CTkFont(weight="bold"), state="disabled", command=lambda: self.execute_operation("PURCHASE_DEBIT"))
        self.btn_purchase_adj.pack(fill="x", padx=15, pady=(4, 15))
        
        self.btn_jump_passbook = ctk.CTkButton(right_pane, text="📋 Open Passbook Statement History", height=35, fg_color="transparent", border_color="#222227", border_width=1, text_color="lightgray", command=self.switch_to_passbook_tab)
        self.btn_jump_passbook.pack(fill="x", padx=20, pady=15)

    def on_search_changed(self):
        self.current_page = 1
        self.fetch_directory_data()

    def go_prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.fetch_directory_data()

    def go_next_page(self):
        query = self.search_entry.get().strip()
        total = db.get_ledger_customers_count(query)
        total_pages = max((total + self.page_size - 1) // self.page_size, 1)
        if self.current_page < total_pages:
            self.current_page += 1
            self.fetch_directory_data()

    def fetch_directory_data(self):
        query = self.search_entry.get().strip()
        total = db.get_ledger_customers_count(query)
        total_pages = max((total + self.page_size - 1) // self.page_size, 1)

        if self.current_page > total_pages:
            self.current_page = total_pages
        if self.current_page < 1:
            self.current_page = 1

        records = db.get_ledger_customers(query, page=self.current_page, page_size=self.page_size)
        
        for w in self.dir_table_scroll.winfo_children(): w.destroy()
        
        # Layout widths mapping row matrix structure
        widths = [60, 130, 95, 95, 32, 32]
        
        h_frame = ctk.CTkFrame(self.dir_table_scroll, fg_color="#16161a", height=32, corner_radius=2)
        h_frame.pack(fill="x", pady=(0, 6))
        
        headers = ["Khata ID", "Account Name", "Phone Ref", "Net Wallet", "", ""]
        for idx, txt in enumerate(headers):
            anchor_pos = "w" if idx == 1 else "center"
            ctk.CTkLabel(h_frame, text=txt, width=widths[idx], font=ctk.CTkFont(size=11, weight="bold"), text_color="#4a90e2", anchor=anchor_pos).pack(side="left", padx=3)

        if not records:
            ctk.CTkLabel(self.dir_table_scroll, text="No wallet accounts found.", text_color="gray", font=ctk.CTkFont(size=12)).pack(pady=30)

        for row_idx, data in enumerate(records):
            k_id, name, phone, balance = data
            
            row = ctk.CTkFrame(self.dir_table_scroll, fg_color="#18181f" if row_idx % 2 == 0 else "transparent", height=40)
            row.pack(fill="x", pady=2)
            
            ctk.CTkLabel(row, text=f"KH-{k_id:04d}", width=widths[0], font=ctk.CTkFont(weight="bold"), text_color="#8ed1fc").pack(side="left", padx=3)
            ctk.CTkLabel(row, text=name[:18], width=widths[1], anchor="w", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=3)
            ctk.CTkLabel(row, text=phone if phone else "—", width=widths[2], text_color="gray").pack(side="left", padx=3)
            
            bal_color = "#4aff4a" if balance >= 0 else "#ff4a4a"
            bal_prefix = "Rs." if balance >= 0 else "-Rs."
            ctk.CTkLabel(row, text=f"{bal_prefix}{abs(balance):,.1f}", width=widths[3], font=ctk.CTkFont(weight="bold"), text_color=bal_color).pack(side="left", padx=3)

            ctk.CTkButton(row, text="🗑️", width=widths[5], height=24, fg_color="#3a1c1c", hover_color="#542424",
                          text_color="#ff8080", font=ctk.CTkFont(size=12),
                          command=lambda p=data: self.confirm_delete_khata(p)).pack(side="right", padx=2)

            ctk.CTkButton(row, text="✏️", width=widths[4], height=24, fg_color="#2b3a4a", hover_color="#3d5268",
                          font=ctk.CTkFont(size=12),
                          command=lambda p=data: self.open_edit_khata_modal(p)).pack(side="right", padx=2)

            ctk.CTkButton(row, text="Focus", width=60, height=24, fg_color="#2b3a4a", font=ctk.CTkFont(size=11), command=lambda p=data: self.focus_wallet_profile(p)).pack(side="right", padx=2)

        # Update pagination controls
        self.lbl_page_info.configure(text=f"Page {self.current_page} of {total_pages}  ({total} total)")
        self.btn_prev_page.configure(state="normal" if self.current_page > 1 else "disabled")
        self.btn_next_page.configure(state="normal" if self.current_page < total_pages else "disabled")

    def focus_wallet_profile(self, payload):
        self.active_customer = payload
        k_id, name, phone, balance = payload
        
        self.lbl_focused_name.configure(text=f"{name} [ID: KH-{k_id:04d}]")
        
        bal_color = "#4aff4a" if balance >= 0 else "#ff4a4a"
        bal_prefix = "Rs. " if balance >= 0 else "-Rs. "
        self.lbl_wallet_balance_tile.configure(text=f"{bal_prefix}{abs(balance):,.2f}", text_color=bal_color)
        
        self.btn_deposit.configure(state="normal")
        self.btn_withdraw.configure(state="normal")
        self.btn_purchase_adj.configure(state="normal")

    def execute_operation(self, txn_type):
        if not self.active_customer: return
        amt_s = self.inp_op_amount.get().strip()
        desc = self.inp_op_desc.get().strip()
        
        if not amt_s:
            messagebox.showerror("Validation Error", "Amount field cannot be left hollow!"); return
        try:
            amt = float(amt_s)
            if amt <= 0: raise ValueError()
        except ValueError:
            messagebox.showerror("Validation Error", "Please parse a valid positive structural numeric entry amount."); return
            
        khata_id = self.active_customer[0]
        final_desc = desc if desc else f"Manual Desk Adjustment Logic ({txn_type})"
        
        success, err = db.process_wallet_transaction(khata_id, txn_type, amt, final_desc)
        if success:
            self.inp_op_amount.delete(0, "end")
            self.inp_op_desc.delete(0, "end")
            
            # Instantly re-pull updated cache lists matrix arrays
            records = db.get_ledger_customers(self.search_entry.get().strip(), page=self.current_page, page_size=self.page_size)
            for updated_node in records:
                if updated_node[0] == khata_id:
                    self.focus_wallet_profile(updated_node)
                    break
            self.fetch_directory_data()
            messagebox.showinfo("Ledger Updated", f"Wallet balance engine synced perfectly for Transaction Type: {txn_type}")
        else:
            messagebox.showerror("Execution Aborted", f"Database rejected log: {err}")

    # =================================================================
    # 📖 SUB-TAB 2: PASSBOOK TIMELINE HISTORY LOGS RUNNING VIEW
    # =================================================================
    def init_passbook_tab_ui(self):
        self.passbook_frame = ctk.CTkFrame(self.view_container, fg_color="transparent")
        
        top_bar = ctk.CTkFrame(self.passbook_frame, fg_color="#16161a", border_color="#222227", border_width=1, height=45)
        top_bar.pack(fill="x", pady=(0, 10))
        
        self.lbl_passbook_title = ctk.CTkLabel(top_bar, text="Statement History Log File Viewer", font=ctk.CTkFont(size=14, weight="bold"), text_color="#8ed1fc")
        self.lbl_passbook_title.pack(side="left", padx=15, pady=8)
        
        ctk.CTkButton(top_bar, text="⬅️ Back To Directory", width=130, fg_color="#121214", border_color="#222227", border_width=1, command=self.switch_to_directory_tab).pack(side="right", padx=15, pady=6)
        
        ctk.CTkButton(top_bar, text="📥 Export This Customer", width=155, height=32,
                      fg_color="#1c3a27", text_color="#4aff4a",
                      border_color="#2d5a3b", border_width=1,
                      font=ctk.CTkFont(size=11, weight="bold"),
                      command=self.export_single_customer_xlsx).pack(side="right", padx=6, pady=6)

        # --- DATE FILTER BAR ---
        filter_bar = ctk.CTkFrame(self.passbook_frame, fg_color="#16161a", border_color="#222227", border_width=1, height=50)
        filter_bar.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(filter_bar, text="From:", font=ctk.CTkFont(size=12), text_color="#8a8a93").pack(side="left", padx=(15, 5), pady=8)
        self.inp_pb_from = ctk.CTkEntry(filter_bar, placeholder_text="YYYY-MM-DD", width=110, height=32, fg_color="#121214", border_color="#222227")
        self.inp_pb_from.pack(side="left", padx=5, pady=8)

        ctk.CTkLabel(filter_bar, text="To:", font=ctk.CTkFont(size=12), text_color="#8a8a93").pack(side="left", padx=(10, 5), pady=8)
        self.inp_pb_to = ctk.CTkEntry(filter_bar, placeholder_text="YYYY-MM-DD", width=110, height=32, fg_color="#121214", border_color="#222227")
        self.inp_pb_to.pack(side="left", padx=5, pady=8)

        ctk.CTkButton(filter_bar, text="🔍 Apply", width=80, height=32, fg_color="#1f293d", font=ctk.CTkFont(size=12, weight="bold"), command=self.apply_pb_date_filter).pack(side="left", padx=(10, 5), pady=8)

        ctk.CTkLabel(filter_bar, text="Month:", font=ctk.CTkFont(size=12), text_color="#8a8a93").pack(side="left", padx=(20, 5), pady=8)

        months = ["All", "January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        self.pb_month_dropdown = ctk.CTkComboBox(filter_bar, values=months, width=130, height=32,
                                                  fg_color="#121214", border_color="#222227",
                                                  button_color="#1f293d", dropdown_fg_color="#1a1a1e",
                                                  command=lambda v: self.apply_month_filter())
        self.pb_month_dropdown.set("All")
        self.pb_month_dropdown.pack(side="left", padx=5, pady=8)

        years = [str(y) for y in range(datetime.date.today().year, datetime.date.today().year - 5, -1)]
        self.pb_year_dropdown = ctk.CTkComboBox(filter_bar, values=years, width=90, height=32,
                                                 fg_color="#121214", border_color="#222227",
                                                 button_color="#1f293d", dropdown_fg_color="#1a1a1e",
                                                 command=lambda v: self.apply_month_filter())
        self.pb_year_dropdown.set(str(datetime.date.today().year))
        self.pb_year_dropdown.pack(side="left", padx=5, pady=8)

        ctk.CTkButton(filter_bar, text="✖ Clear", width=80, height=32, fg_color="transparent", border_color="#444", border_width=1, text_color="#a0a0a9", font=ctk.CTkFont(size=12), command=self.clear_passbook_filter).pack(side="left", padx=(10, 5), pady=8)

        self.passbook_scroll = ctk.CTkScrollableFrame(self.passbook_frame, fg_color="#121214", border_color="#222227", border_width=1, corner_radius=6)
        self.passbook_scroll.pack(fill="both", expand=True)

        # --- PASSBOOK PAGINATION BAR ---
        pb_pagination_bar = ctk.CTkFrame(self.passbook_frame, fg_color="#16161a", border_color="#222227", border_width=1, height=45)
        pb_pagination_bar.pack(fill="x", pady=(10, 0))

        self.btn_pb_prev_page = ctk.CTkButton(pb_pagination_bar, text="⬅ Previous", width=100, height=32, fg_color="#1f1f24", font=ctk.CTkFont(size=12), command=self.go_pb_prev_page)
        self.btn_pb_prev_page.pack(side="left", padx=10, pady=6)

        self.lbl_pb_page_info = ctk.CTkLabel(pb_pagination_bar, text="Page 1 of 1", font=ctk.CTkFont(size=12, weight="bold"), text_color="#a0a0a9")
        self.lbl_pb_page_info.pack(side="left", expand=True)

        self.btn_pb_next_page = ctk.CTkButton(pb_pagination_bar, text="Next ➡", width=100, height=32, fg_color="#1f1f24", font=ctk.CTkFont(size=12), command=self.go_pb_next_page)
        self.btn_pb_next_page.pack(side="right", padx=10, pady=6)

    def apply_pb_date_filter(self):
        self.pb_current_page = 1
        self.render_passbook_timeline()

    def go_pb_prev_page(self):
        if self.pb_current_page > 1:
            self.pb_current_page -= 1
            self.render_passbook_timeline()

    def go_pb_next_page(self):
        if not self.active_customer:
            return
        k_id, _, _, _ = self.active_customer
        start_date = self.inp_pb_from.get().strip()
        end_date = self.inp_pb_to.get().strip()
        total = db.get_customer_passbook_count(k_id, start_date, end_date)
        total_pages = max((total + self.pb_page_size - 1) // self.pb_page_size, 1)
        if self.pb_current_page < total_pages:
            self.pb_current_page += 1
            self.render_passbook_timeline()

    def apply_month_filter(self):
        month_name = self.pb_month_dropdown.get()
        year = self.pb_year_dropdown.get()
        self.pb_current_page = 1

        if month_name == "All":
            self.inp_pb_from.delete(0, "end")
            self.inp_pb_to.delete(0, "end")
            self.render_passbook_timeline()
            return

        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        month_num = months.index(month_name) + 1

        first_day = datetime.date(int(year), month_num, 1)
        if month_num == 12:
            last_day = datetime.date(int(year), 12, 31)
        else:
            last_day = datetime.date(int(year), month_num + 1, 1) - datetime.timedelta(days=1)

        self.inp_pb_from.delete(0, "end")
        self.inp_pb_from.insert(0, first_day.strftime("%Y-%m-%d"))
        self.inp_pb_to.delete(0, "end")
        self.inp_pb_to.insert(0, last_day.strftime("%Y-%m-%d"))

        self.render_passbook_timeline()

    def clear_passbook_filter(self):
        self.inp_pb_from.delete(0, "end")
        self.inp_pb_to.delete(0, "end")
        self.pb_month_dropdown.set("All")
        self.pb_current_page = 1
        self.render_passbook_timeline()

    def render_passbook_timeline(self):
        if not self.active_customer: return
        k_id, name, _, _ = self.active_customer
        
        self.lbl_passbook_title.configure(text=f"📖 Ledger Passbook Statement Audit Timeline: {name} [KH-{k_id:04d}]")
        for w in self.passbook_scroll.winfo_children(): w.destroy()
        
        start_date = self.inp_pb_from.get().strip()
        end_date = self.inp_pb_to.get().strip()

        total = db.get_customer_passbook_count(k_id, start_date, end_date)
        total_pages = max((total + self.pb_page_size - 1) // self.pb_page_size, 1)

        if self.pb_current_page > total_pages:
            self.pb_current_page = total_pages
        if self.pb_current_page < 1:
            self.pb_current_page = 1

        logs = db.get_customer_passbook(k_id, start_date, end_date, page=self.pb_current_page, page_size=self.pb_page_size)
        widths = [130, 130, 180, 85, 95]
        
        h_frame = ctk.CTkFrame(self.passbook_scroll, fg_color="#16161a", height=32, corner_radius=2)
        h_frame.pack(fill="x", pady=(0, 6))
        
        headers = ["Timestamp Date", "Action Type Log", "Narration Details", "Amount", "Closing Balance"]
        for idx, txt in enumerate(headers):
            anchor_pos = "w" if idx == 2 else "center"
            ctk.CTkLabel(h_frame, text=txt, width=widths[idx], font=ctk.CTkFont(size=11, weight="bold"), text_color="#4a90e2", anchor=anchor_pos).pack(side="left", padx=4)

        if not logs:
            ctk.CTkLabel(self.passbook_scroll, text="Is date range mein koi transaction nahi mili.", text_color="gray", font=ctk.CTkFont(size=12)).pack(pady=30)
            
        for row_idx, item in enumerate(logs):
            ts, action, desc, amt, closing = item
            row = ctk.CTkFrame(self.passbook_scroll, fg_color="#18181f" if row_idx % 2 == 0 else "transparent", height=38)
            row.pack(fill="x", pady=2)
            
            ctk.CTkLabel(row, text=ts, width=widths[0], text_color="gray", font=ctk.CTkFont(size=11)).pack(side="left", padx=4)
            
            # Action Colors Parsing Matrix Tags
            act_colors = {'ADVANCE_DEPOSIT': '#4aff4a', 'CASH_WITHDRAWAL': '#ff9f43', 'PURCHASE_DEBIT': '#8ed1fc', 'OVERDRAFT_CREDIT': '#ff4a4a'}
            ctk.CTkLabel(row, text=action, width=widths[1], text_color=act_colors.get(action, '#ffffff'), font=ctk.CTkFont(size=11, weight="bold")).pack(side="left", padx=4)
            
            ctk.CTkLabel(row, text=desc[:28], width=widths[2], anchor="w", font=ctk.CTkFont(size=11)).pack(side="left", padx=4)
            ctk.CTkLabel(row, text=f"{amt:,.1f}", width=widths[3], font=ctk.CTkFont(size=11)).pack(side="left", padx=4)
            
            c_color = "#4aff4a" if closing >= 0 else "#ff4a4a"
            c_prefix = "Rs." if closing >= 0 else "-Rs."
            ctk.CTkLabel(row, text=f"{c_prefix}{abs(closing):,.1f}", width=widths[4], font=ctk.CTkFont(size=11, weight="bold"), text_color=c_color).pack(side="left", padx=4)

        # Update pagination controls
        self.lbl_pb_page_info.configure(text=f"Page {self.pb_current_page} of {total_pages}  ({total} total)")
        self.btn_pb_prev_page.configure(state="normal" if self.pb_current_page > 1 else "disabled")
        self.btn_pb_next_page.configure(state="normal" if self.pb_current_page < total_pages else "disabled")

    # =================================================================
    # 📊 SUB-TAB 3: AGENCY DAILY CASH FLOW METRICS REPORT SUMMARY
    # =================================================================
    def init_cashflow_tab_ui(self):
        self.cashflow_frame = ctk.CTkFrame(self.view_container, fg_color="transparent")
        
        card = ctk.CTkFrame(self.cashflow_frame, fg_color="#16161a", border_color="#222227", border_width=1, corner_radius=6)
        card.pack(padx=40, pady=40, fill="both", expand=True)
        
        ctk.CTkLabel(card, text="📊 Live Agency Cash Flow Dashboard Summary (Today)", font=ctk.CTkFont(size=16, weight="bold"), text_color="#4a90e2").pack(pady=25)
        
        # Inwards Advance Tile
        self.lbl_flow_in = ctk.CTkLabel(card, text="Total Advance Cash Deposited In: Rs. 0.00", font=ctk.CTkFont(size=15, weight="bold"), text_color="#4aff4a")
        self.lbl_flow_in.pack(pady=15)
        
        # Outwards Advance Tile
        self.lbl_flow_out = ctk.CTkLabel(card, text="Total Advance Cash Withdrawn Out: Rs. 0.00", font=ctk.CTkFont(size=15, weight="bold"), text_color="#ff9f43")
        self.lbl_flow_out.pack(pady=15)
        
        # Separation Bar
        ctk.CTkFrame(card, height=1, fg_color="#222227").pack(fill="x", padx=100, pady=10)
        
        # Net Vault State Balance Tracker Card
        self.lbl_flow_net = ctk.CTkLabel(card, text="Net Wallet Balance Vault Flow: Rs. 0.00", font=ctk.CTkFont(size=18, weight="bold"), text_color="#ffffff")
        self.lbl_flow_net.pack(pady=20)
        
        ctk.CTkButton(card, text="🔄 Refresh Calculations Data Engine", height=38, fg_color="#1f293d", font=ctk.CTkFont(weight="bold"), command=self.render_cashflow_metrics).pack(pady=25)

    def render_cashflow_metrics(self):
        cash_in, cash_out = db.get_daily_cashflow_summary()
        net_vault = cash_in - cash_out
        
        self.lbl_flow_in.configure(text=f"Total Advance Cash Deposited In: Rs. {cash_in:,.2f}")
        self.lbl_flow_out.configure(text=f"Total Advance Cash Withdrawn Out: Rs. {cash_out:,.2f}")
        
        nv_color = "#4aff4a" if net_vault >= 0 else "#ff4a4a"
        nv_prefix = "Rs. " if net_vault >= 0 else "-Rs. "
        self.lbl_flow_net.configure(text=f"Net Wallet Balance Vault Flow: {nv_prefix}{abs(net_vault):,.2f}", text_color=nv_color)

    # =================================================================
    # 📑 REGISTER WALLET CUSTOMER ACCOUNT MODAL WINDOW CONTROL
    # =================================================================
    def open_new_wallet_modal(self):
        modal = ctk.CTkToplevel(self)
        modal.title("Open Customer Advance Wallet Profile")
        modal.geometry("400x420")
        modal.resizable(False, False)
        modal.attributes("-topmost", True)
        modal.grab_set()
        modal.configure(fg_color="#121214")
        
        w = ctk.CTkFrame(modal, fg_color="#1a1a1e", border_color="#222227", border_width=1, corner_radius=6)
        w.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(w, text="Open New Advance Wallet", font=ctk.CTkFont(size=15, weight="bold"), text_color="#4a90e2").pack(pady=(15, 10))
        
        inp_n = ctk.CTkEntry(w, placeholder_text="Customer Full Name *", width=300, height=38, fg_color="#121214", border_color="#222227")
        inp_n.pack(pady=8)
        
        inp_p = ctk.CTkEntry(w, placeholder_text="Phone Number (optional)", width=300, height=38, fg_color="#121214", border_color="#222227")
        inp_p.pack(pady=8)

        ctk.CTkLabel(w, text="Previous Balance / Udhaar (optional)", font=ctk.CTkFont(size=11), text_color="#8a8a93").pack(anchor="w", padx=50, pady=(6, 0))
        inp_bal = ctk.CTkEntry(w, placeholder_text="e.g. -5000 (udhaar) or 2000 (advance)", width=300, height=38, fg_color="#121214", border_color="#222227")
        inp_bal.pack(pady=8)
        ctk.CTkLabel(w, text="Negative = customer owes us (udhaar). Positive = customer advance.",
                     font=ctk.CTkFont(size=10), text_color="#6e6e77", wraplength=300, justify="left").pack(anchor="w", padx=50, pady=(0, 4))
        
        lbl_err = ctk.CTkLabel(w, text="", font=ctk.CTkFont(size=11))
        lbl_err.pack(pady=2)
        
        def save():
            name, phone = inp_n.get().strip(), inp_p.get().strip()
            bal_raw = inp_bal.get().strip()

            if not name:
                lbl_err.configure(text="Error: Customer name is mandatory.", text_color="#ff4a4a"); return

            try:
                opening_balance = float(bal_raw) if bal_raw else 0.0
            except ValueError:
                lbl_err.configure(text="Error: Previous balance must be a number.", text_color="#ff4a4a"); return

            success, msg = db.add_ledger_customer(name, phone, opening_balance)
            if success:
                modal.destroy()
                self.fetch_directory_data()
                messagebox.showinfo("Wallet Created", "Customer identity successfully mapped with unique system Khata ID.")
            else:
                lbl_err.configure(text=f"Rejected: {msg}", text_color="#ff4a4a")
                
        ctk.CTkButton(w, text="🚀 Open Secure Wallet Account File", height=40, fg_color="#1f293d", font=ctk.CTkFont(weight="bold"), command=save).pack(fill="x", padx=30, pady=15)

    # ── EDIT: Khata customer name/phone ─────────────────────────────
    def open_edit_khata_modal(self, payload):
        k_id, name, phone, balance = payload

        modal = ctk.CTkToplevel(self)
        modal.title("Edit Customer Account")
        modal.geometry("400x300")
        modal.resizable(False, False)
        modal.attributes("-topmost", True)
        modal.grab_set()
        modal.configure(fg_color="#121214")

        w = ctk.CTkFrame(modal, fg_color="#1a1a1e", border_color="#222227", border_width=1, corner_radius=6)
        w.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(w, text=f"Edit Account - KH-{k_id:04d}", font=ctk.CTkFont(size=15, weight="bold"), text_color="#4a90e2").pack(pady=(15, 10))

        inp_n = ctk.CTkEntry(w, placeholder_text="Account Customer Full Name *", width=300, height=38, fg_color="#121214", border_color="#222227")
        inp_n.insert(0, name)
        inp_n.pack(pady=8)

        inp_p = ctk.CTkEntry(w, placeholder_text="Phone Number (optional)", width=300, height=38, fg_color="#121214", border_color="#222227")
        inp_p.insert(0, phone if phone else "")
        inp_p.pack(pady=8)

        lbl_err = ctk.CTkLabel(w, text="", font=ctk.CTkFont(size=11))
        lbl_err.pack(pady=2)

        def save():
            new_name, new_phone = inp_n.get().strip(), inp_p.get().strip()
            if not new_name:
                lbl_err.configure(text="Error: Customer name is mandatory.", text_color="#ff4a4a"); return

            success, msg = db.update_ledger_customer(k_id, new_name, new_phone)
            if success:
                modal.destroy()
                self.fetch_directory_data()
                messagebox.showinfo("Updated", "Account details updated successfully.")
            else:
                lbl_err.configure(text=f"Rejected: {msg}", text_color="#ff4a4a")

        ctk.CTkButton(w, text="💾 Save Changes", height=40, fg_color="#1f293d", font=ctk.CTkFont(weight="bold"), command=save).pack(fill="x", padx=30, pady=15)

    # ── DELETE: Khata customer ───────────────────────────────────────
    def confirm_delete_khata(self, payload):
        k_id, name, phone, balance = payload
        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Delete account 'KH-{k_id:04d} - {name}'?\n\n"
            "This will permanently remove the account and its full transaction history.\n\n"
            "This action cannot be undone."
        )
        if not confirm:
            return

        success, msg = db.delete_ledger_customer(k_id)
        if success:
            self.fetch_directory_data()
            messagebox.showinfo("Deleted", msg)
        else:
            messagebox.showerror("Delete Failed", msg)

    # ── IMPORT: Khata customers from Excel sheet ─────────────────────
    def import_khata_from_xlsx(self):
        if not XLSX_AVAILABLE:
            messagebox.showerror("Missing Library", "openpyxl install nahi hai.\npip install openpyxl")
            return

        file_path = filedialog.askopenfilename(
            title="Import Khata Sheet",
            filetypes=[("Excel File", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            rows_data = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue

                name_cell, phone_cell, balance_cell = (row + (None, None, None))[:3]

                # Skip header row if first column isn't usable as a name (e.g. literally "Name")
                if isinstance(name_cell, str) and name_cell.strip().lower() in ("name", "customer name", "account name"):
                    continue

                if name_cell is None and phone_cell is None:
                    continue

                rows_data.append((
                    name_cell if name_cell is not None else "",
                    phone_cell if phone_cell is not None else "",
                    balance_cell if balance_cell is not None else 0
                ))

            if not rows_data:
                messagebox.showwarning("Empty File", "Is sheet mein koi valid record nahi mila.\n\nExpected columns: Name, Phone, Opening Balance")
                return

            success_count, skipped_count, errors = db.import_ledger_customers_bulk(rows_data)

            self.current_page = 1
            self.fetch_directory_data()

            summary = (
                f"Import Mukammal Ho Gaya ✅\n\n"
                f"• Naye Khata Accounts: {success_count}\n"
                f"• Skip Hue (duplicate phone / khali row): {skipped_count}"
            )
            if errors:
                summary += f"\n\n⚠️ {len(errors)} error(s) occurred during import."

            messagebox.showinfo("Import Successful", summary)

        except Exception as e:
            messagebox.showerror("Import Failed", str(e))

    # =================================================================
    # 📊 EXPORT FUNCTIONS
    # =================================================================

    def _make_xlsx_styles(self):
        """Shared style objects for all exports"""
        return {
            "hdr_fill":    PatternFill("solid", fgColor="1e3a5f"),
            "hdr_font":    Font(color="FFFFFF", bold=True, size=11, name="Arial"),
            "title_font":  Font(color="FFFFFF", bold=True, size=14, name="Arial"),
            "green_font":  Font(color="4aff4a", bold=True, name="Arial"),
            "red_font":    Font(color="ff4a4a", bold=True, name="Arial"),
            "orange_font": Font(color="ff9f43", bold=True, name="Arial"),
            "blue_font":   Font(color="8ed1fc", bold=True, name="Arial"),
            "muted_font":  Font(color="94a3b8", name="Arial", size=10),
            "body_font":   Font(color="f1f5f9", name="Arial", size=10),
            "center":      Alignment(horizontal="center", vertical="center"),
            "right":       Alignment(horizontal="right"),
            "border":      Border(
                left=Side(style="thin", color="2a2e38"),
                right=Side(style="thin", color="2a2e38"),
                top=Side(style="thin", color="2a2e38"),
                bottom=Side(style="thin", color="2a2e38")
            ),
            "dark_fill":   PatternFill("solid", fgColor="1a1d24"),
            "darker_fill": PatternFill("solid", fgColor="15171d"),
            "title_fill":  PatternFill("solid", fgColor="0f1117"),
        }

    def _write_passbook_sheet(self, ws, khata_id, cust_name, phone, balance, s):
        """Passbook timeline sheet writer — reused for single & all exports"""
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = f"📖 Passbook: {cust_name}  |  KH-{khata_id:04d}  |  {phone}"
        c.font = s["title_font"]; c.fill = s["title_fill"]; c.alignment = s["center"]
        ws.row_dimensions[1].height = 28

        # Balance info row
        ws.merge_cells("A2:F2")
        c = ws["A2"]
        bal_str = f"Rs. {balance:,.2f}" if balance >= 0 else f"-Rs. {abs(balance):,.2f}"
        c.value = f"Current Wallet Balance: {bal_str}   |   Exported: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        c.font = s["muted_font"]; c.fill = PatternFill("solid", fgColor="11131a"); c.alignment = s["center"]
        ws.row_dimensions[2].height = 18

        # Headers
        hdrs = ["#", "Timestamp", "Action Type", "Narration / Description", "Amount (Rs.)", "Closing Balance (Rs.)"]
        ws.row_dimensions[3].height = 22
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = s["hdr_font"]; c.fill = s["hdr_fill"]
            c.alignment = s["center"]; c.border = s["border"]

        # Data
        logs = db.get_customer_passbook(khata_id, page=1, page_size=10**9)
        action_colors = {
            "ADVANCE_DEPOSIT":  "4aff4a",
            "CASH_WITHDRAWAL":  "ff9f43",
            "PURCHASE_DEBIT":   "8ed1fc",
            "OVERDRAFT_CREDIT": "ff4a4a",
        }
        for ri, (ts, action, desc, amt, closing) in enumerate(logs, 4):
            fill = s["dark_fill"] if ri % 2 == 0 else s["darker_fill"]
            row_vals = [len(logs) - ri + 4, ts, action, desc, round(amt, 2), round(closing, 2)]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill; c.border = s["border"]
                if ci == 3:
                    clr = action_colors.get(action, "f1f5f9")
                    c.font = Font(color=clr, bold=True, name="Arial", size=10)
                elif ci in (5, 6):
                    c.number_format = '#,##0.00'
                    c.alignment = s["right"]
                    if ci == 6:
                        c.font = Font(color="4aff4a" if closing >= 0 else "ff4a4a", bold=True, name="Arial", size=10)
                    else:
                        c.font = s["body_font"]
                else:
                    c.font = s["body_font"]

        col_widths = [5, 20, 18, 38, 16, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        return len(logs)

    def _write_invoices_sheet(self, ws, cust_name, s):
        """Invoice history sheet for a customer"""
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = f"🧾 Invoice History: {cust_name}"
        c.font = s["title_font"]; c.fill = s["title_fill"]; c.alignment = s["center"]
        ws.row_dimensions[1].height = 28

        hdrs = ["Invoice No", "Date & Time", "Payment Mode", "Subtotal (Rs.)", "Tax %", "Net Amount (Rs.)", "Profit (Rs.)", "Items Detail"]
        ws.row_dimensions[2].height = 22
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = s["hdr_font"]; c.fill = s["hdr_fill"]
            c.alignment = s["center"]; c.border = s["border"]

        invoices = db.get_customer_linked_invoices(cust_name)
        current_row = 3

        for inv in invoices:
            inv_id, c_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit = inv
            items = db.get_invoice_profit(inv_id)
            items_str = "; ".join([f"{it[0]} x{it[1]}" for it in items]) if items else "—"

            fill = s["dark_fill"] if current_row % 2 == 0 else s["darker_fill"]
            row_vals = [
                f"INV-{inv_id:04d}", timestamp, pay_mode,
                round(subtotal, 2), f"{tax_pct}%",
                round(net_amount, 2),
                round(profit, 2) if profit is not None else "—",
                items_str
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.fill = fill; c.border = s["border"]
                c.font = s["body_font"]
                if ci in (4, 6):
                    c.number_format = '#,##0.00'; c.alignment = s["right"]
                elif ci == 7 and profit is not None:
                    c.font = Font(color="a855f7", bold=True, name="Arial", size=10)
                    c.number_format = '#,##0.00'; c.alignment = s["right"]
                elif ci in (1, 3, 5):
                    c.alignment = s["center"]
            current_row += 1

        # Total footer
        if invoices:
            current_row += 1
            total_net = sum(inv[4] for inv in invoices)
            profits = [inv[7] for inv in invoices if inv[7] is not None]
            ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True, color="94a3b8", name="Arial")
            c6 = ws.cell(row=current_row, column=6, value=round(total_net, 2))
            c6.font = Font(color="22c55e", bold=True, name="Arial")
            c6.number_format = '#,##0.00'; c6.alignment = s["right"]
            if profits:
                c7 = ws.cell(row=current_row, column=7, value=round(sum(profits), 2))
                c7.font = Font(color="a855f7", bold=True, name="Arial")
                c7.number_format = '#,##0.00'; c7.alignment = s["right"]

        col_widths = [13, 20, 14, 16, 8, 18, 14, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── EXPORT: Single Customer ──────────────────────────────────────
    def export_single_customer_xlsx(self):
        if not self.active_customer:
            messagebox.showwarning("No Customer", "Pehle directory se kisi customer ko focus karein.")
            return

        k_id, name, phone, balance = self.active_customer

        if not XLSX_AVAILABLE:
            messagebox.showerror("Missing Library", "openpyxl install nahi hai.\npip install openpyxl")
            return

        safe_name = "".join(c for c in name if c.isalnum() or c in " _-")[:20].strip()
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")],
            initialfile=f"Khata_{safe_name}_{datetime.date.today()}.xlsx",
            title="Export Customer Khata"
        )
        if not save_path: return

        try:
            wb = openpyxl.Workbook()
            s = self._make_xlsx_styles()

            # Sheet 1: Passbook Timeline
            ws1 = wb.active
            ws1.title = "Passbook Timeline"
            self._write_passbook_sheet(ws1, k_id, name, phone, balance, s)

            # Sheet 2: Invoice History
            ws2 = wb.create_sheet("Invoice History")
            self._write_invoices_sheet(ws2, name, s)

            wb.save(save_path)
            messagebox.showinfo(
                "Export Successful ✅",
                f"{name} ka pura khata export ho gaya!\n\n"
                f"• Sheet 1: Passbook Timeline (deposits, withdrawals, purchases)\n"
                f"• Sheet 2: Invoice History (saari purchases detail ke saath)\n\n"
                f"File:\n{save_path}"
            )
            try: os.startfile(save_path)
            except Exception: pass

        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    # ── EXPORT: All Customers (one sheet per customer) ───────────────
    def export_all_khata_to_xlsx(self):
        if not XLSX_AVAILABLE:
            messagebox.showerror("Missing Library", "openpyxl install nahi hai.\npip install openpyxl")
            return

        all_customers = db.get_all_ledger_customers_full()
        if not all_customers:
            messagebox.showwarning("No Data", "Koi bhi khata customer register nahi hai abhi.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")],
            initialfile=f"AllKhata_Export_{datetime.date.today()}.xlsx",
            title="Export All Khata Accounts"
        )
        if not save_path: return

        try:
            wb = openpyxl.Workbook()
            s = self._make_xlsx_styles()

            # ── Sheet 1: Master Directory ──────────────────────────
            ws_dir = wb.active
            ws_dir.title = "Master Directory"
            ws_dir.sheet_view.showGridLines = False

            ws_dir.merge_cells("A1:F1")
            c = ws_dir["A1"]
            c.value = f"💳 Khata Master Directory — {datetime.date.today()}"
            c.font = s["title_font"]; c.fill = s["title_fill"]; c.alignment = s["center"]
            ws_dir.row_dimensions[1].height = 28

            dir_hdrs = ["Khata ID", "Customer Name", "Phone", "Wallet Balance (Rs.)", "Account Opened", "Status"]
            ws_dir.row_dimensions[2].height = 22
            for ci, h in enumerate(dir_hdrs, 1):
                c = ws_dir.cell(row=2, column=ci, value=h)
                c.font = s["hdr_font"]; c.fill = s["hdr_fill"]
                c.alignment = s["center"]; c.border = s["border"]

            for ri, cust in enumerate(all_customers, 3):
                k_id, cust_name, phone, balance, created_at = cust
                fill = s["dark_fill"] if ri % 2 == 0 else s["darker_fill"]
                status = "✅ Positive" if balance >= 0 else "🔴 Udhaar"
                row_vals = [f"KH-{k_id:04d}", cust_name, phone, round(balance, 2), created_at, status]
                for ci, val in enumerate(row_vals, 1):
                    c = ws_dir.cell(row=ri, column=ci, value=val)
                    c.fill = fill; c.border = s["border"]
                    if ci == 4:
                        c.number_format = '#,##0.00'; c.alignment = s["right"]
                        c.font = Font(color="4aff4a" if balance >= 0 else "ff4a4a", bold=True, name="Arial", size=10)
                    elif ci == 6:
                        c.font = Font(color="4aff4a" if balance >= 0 else "ff4a4a", name="Arial", size=10)
                        c.alignment = s["center"]
                    else:
                        c.font = s["body_font"]

            # ── Totals row ──────────────────────────────────────────
            total_advance = sum(c[3] for c in all_customers if c[3] > 0)
            total_udhaar = sum(c[3] for c in all_customers if c[3] < 0)
            net_total = total_advance + total_udhaar

            totals_row_idx = 3 + len(all_customers) + 1
            ws_dir.merge_cells(start_row=totals_row_idx, start_column=1, end_row=totals_row_idx, end_column=3)
            c = ws_dir.cell(row=totals_row_idx, column=1, value="TOTAL (Net Wallet — All Khata)")
            c.font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
            c.fill = PatternFill("solid", fgColor="1e3a5f")
            c.alignment = s["right"]; c.border = s["border"]
            for col in (2, 3):
                ws_dir.cell(row=totals_row_idx, column=col).fill = PatternFill("solid", fgColor="1e3a5f")
                ws_dir.cell(row=totals_row_idx, column=col).border = s["border"]

            c = ws_dir.cell(row=totals_row_idx, column=4, value=round(net_total, 2))
            c.number_format = '#,##0.00'; c.alignment = s["right"]; c.border = s["border"]
            c.fill = PatternFill("solid", fgColor="1e3a5f")
            c.font = Font(color="4aff4a" if net_total >= 0 else "ff4a4a", bold=True, name="Arial", size=11)
            for col in (5, 6):
                ws_dir.cell(row=totals_row_idx, column=col).fill = PatternFill("solid", fgColor="1e3a5f")
                ws_dir.cell(row=totals_row_idx, column=col).border = s["border"]

            # Total Advance (Credit) row
            adv_row_idx = totals_row_idx + 1
            ws_dir.merge_cells(start_row=adv_row_idx, start_column=1, end_row=adv_row_idx, end_column=3)
            c = ws_dir.cell(row=adv_row_idx, column=1, value="Total Advance (Credit / Positive)")
            c.font = s["body_font"]; c.alignment = s["right"]; c.border = s["border"]
            c.fill = s["dark_fill"]
            for col in (2, 3):
                ws_dir.cell(row=adv_row_idx, column=col).fill = s["dark_fill"]
                ws_dir.cell(row=adv_row_idx, column=col).border = s["border"]
            c = ws_dir.cell(row=adv_row_idx, column=4, value=round(total_advance, 2))
            c.number_format = '#,##0.00'; c.alignment = s["right"]; c.border = s["border"]
            c.fill = s["dark_fill"]
            c.font = Font(color="4aff4a", bold=True, name="Arial", size=10)
            for col in (5, 6):
                ws_dir.cell(row=adv_row_idx, column=col).fill = s["dark_fill"]
                ws_dir.cell(row=adv_row_idx, column=col).border = s["border"]

            # Total Udhaar (Debit) row
            udhaar_row_idx = adv_row_idx + 1
            ws_dir.merge_cells(start_row=udhaar_row_idx, start_column=1, end_row=udhaar_row_idx, end_column=3)
            c = ws_dir.cell(row=udhaar_row_idx, column=1, value="Total Udhaar (Debit / Negative)")
            c.font = s["body_font"]; c.alignment = s["right"]; c.border = s["border"]
            c.fill = s["darker_fill"]
            for col in (2, 3):
                ws_dir.cell(row=udhaar_row_idx, column=col).fill = s["darker_fill"]
                ws_dir.cell(row=udhaar_row_idx, column=col).border = s["border"]
            c = ws_dir.cell(row=udhaar_row_idx, column=4, value=round(total_udhaar, 2))
            c.number_format = '#,##0.00'; c.alignment = s["right"]; c.border = s["border"]
            c.fill = s["darker_fill"]
            c.font = Font(color="ff4a4a", bold=True, name="Arial", size=10)
            for col in (5, 6):
                ws_dir.cell(row=udhaar_row_idx, column=col).fill = s["darker_fill"]
                ws_dir.cell(row=udhaar_row_idx, column=col).border = s["border"]

            # Total Khata Accounts row
            count_row_idx = udhaar_row_idx + 1
            ws_dir.merge_cells(start_row=count_row_idx, start_column=1, end_row=count_row_idx, end_column=3)
            c = ws_dir.cell(row=count_row_idx, column=1, value="Total Khata Accounts")
            c.font = s["body_font"]; c.alignment = s["right"]; c.border = s["border"]
            c.fill = s["dark_fill"]
            for col in (2, 3):
                ws_dir.cell(row=count_row_idx, column=col).fill = s["dark_fill"]
                ws_dir.cell(row=count_row_idx, column=col).border = s["border"]
            c = ws_dir.cell(row=count_row_idx, column=4, value=len(all_customers))
            c.alignment = s["right"]; c.border = s["border"]
            c.fill = s["dark_fill"]
            c.font = Font(color="f1f5f9", bold=True, name="Arial", size=10)
            for col in (5, 6):
                ws_dir.cell(row=count_row_idx, column=col).fill = s["dark_fill"]
                ws_dir.cell(row=count_row_idx, column=col).border = s["border"]

            dir_col_widths = [12, 25, 16, 20, 20, 14]
            for i, w in enumerate(dir_col_widths, 1):
                ws_dir.column_dimensions[get_column_letter(i)].width = w

            # ── Per-customer sheets ────────────────────────────────
            for cust in all_customers:
                k_id, cust_name, phone, balance, created_at = cust
                safe = "".join(c for c in cust_name if c.isalnum() or c == " ")[:20].strip()

                # Passbook sheet
                ws_pb = wb.create_sheet(f"{safe[:15]}_PB")
                self._write_passbook_sheet(ws_pb, k_id, cust_name, phone, balance, s)

                # Invoice sheet
                ws_inv = wb.create_sheet(f"{safe[:14]}_INV")
                self._write_invoices_sheet(ws_inv, cust_name, s)

            wb.save(save_path)
            messagebox.showinfo(
                "Export Successful ✅",
                f"Sab {len(all_customers)} khata accounts export ho gaye!\n\n"
                f"• Sheet 1: Master Directory (sab customers)\n"
                f"• Har customer ke liye 2 sheets:\n"
                f"  – Passbook Timeline (deposits / purchases / withdrawals)\n"
                f"  – Invoice History (kya kharida, invoice items, profit)\n\n"
                f"File:\n{save_path}"
            )
            try: os.startfile(save_path)
            except Exception: pass

        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    def export_all_khata_to_pdf(self):
        """Exports the Khata Master Directory (with totals) as a printable PDF."""
        from modules.pdf_export import build_table_html, export_html_to_pdf

        all_customers = db.get_all_ledger_customers_full()
        if not all_customers:
            messagebox.showwarning("No Data", "Koi bhi khata customer register nahi hai abhi.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF File", "*.pdf")],
            initialfile=f"AllKhata_Export_{datetime.date.today()}.pdf",
            title="Export All Khata Accounts (PDF)"
        )
        if not save_path: return

        try:
            total_advance = sum(c[3] for c in all_customers if c[3] > 0)
            total_udhaar = sum(c[3] for c in all_customers if c[3] < 0)
            net_total = total_advance + total_udhaar

            headers = ["Khata ID", "Customer Name", "Phone", "Wallet Balance (Rs.)", "Account Opened", "Status"]
            col_classes = ["center", "", "center", "right", "center", "center"]

            rows = []
            for cust in all_customers:
                k_id, cust_name, phone, balance, created_at = cust
                bal_text = f"{balance:,.2f}"
                bal_cls = "green" if balance >= 0 else "red"
                status = "Positive" if balance >= 0 else "Udhaar"
                status_cls = "green" if balance >= 0 else "red"
                rows.append([
                    f"KH-{k_id:04d}",
                    cust_name,
                    phone if phone else "—",
                    (bal_text, bal_cls),
                    created_at,
                    (status, status_cls),
                ])

            net_cls = "green" if net_total >= 0 else "red"
            totals_row = [
                ("TOTAL (Net Wallet)", "bold"), "", "",
                (f"{net_total:,.2f}", f"bold {net_cls}"),
                "", ""
            ]

            summary_rows = [
                ("Total Khata Accounts:", f"{len(all_customers)}"),
                ("Total Advance (Credit / Positive):", f"Rs. {total_advance:,.2f}"),
                ("Total Udhaar (Debit / Negative):", f"-Rs. {abs(total_udhaar):,.2f}"),
                ("Net Wallet (All Khata):", f"Rs. {net_total:,.2f}" if net_total >= 0 else f"-Rs. {abs(net_total):,.2f}"),
            ]

            html = build_table_html(
                title="Khata Master Directory",
                subtitle=f"Generated on {datetime.date.today()}",
                headers=headers,
                rows=rows,
                col_classes=col_classes,
                orientation="auto",
                summary_rows=summary_rows,
                totals_row=totals_row,
            )

            success, msg = export_html_to_pdf(html, save_path)
            if success:
                messagebox.showinfo("Export Successful ✅", f"Khata directory PDF saved!\n\n{msg}")
                try: os.startfile(save_path)
                except Exception: pass
            else:
                messagebox.showerror("PDF Export Failed", msg)

        except Exception as e:
            messagebox.showerror("Export Failed", str(e))