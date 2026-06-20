"""
Product Daily Report Module
-----------------------------
Shows a day-by-day breakdown for every product:
    Date | Product | Available | Sold | Remaining | Revenue | Profit

Filters:
    - Custom date range (From / To)
    - Quick presets: This Month, Last 3 Months, Last 6 Months, This Year
    - Year selector (e.g. 2025, 2026 -> full calendar year)

Export:
    - Download as Excel (.xlsx)
"""

import customtkinter as ctk
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
from tkcalendar import DateEntry
import datetime

from modules.product_reports_engine import (
    get_product_daily_report, get_period_presets, get_available_years, get_year_range
)
from modules.pdf_export import build_table_html, export_html_to_pdf

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# =====================================================================
# THEME PALETTE (matches rest of app)
# =====================================================================
COL_BG_DARK     = "#0f1117"
COL_BG_CARD     = "#1a1d24"
COL_BG_CARD_ALT = "#15171d"
COL_BG_INPUT    = "#11131a"
COL_BORDER      = "#2a2e38"
COL_ACCENT      = "#3b82f6"
COL_SUCCESS     = "#22c55e"
COL_DANGER      = "#ef4444"
COL_TEXT_MAIN   = "#f1f5f9"
COL_TEXT_MUTED  = "#94a3b8"


class ProductDailyReportView(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")

        self.cached_rows = []
        self.current_page = 0
        self.page_size = 10

        today = datetime.date.today()
        month_start = today.replace(day=1)

        # =================================================================
        # FILTER CARD
        # =================================================================
        filter_card = ctk.CTkFrame(self, fg_color=COL_BG_CARD,
                                    border_color=COL_BORDER, border_width=1, corner_radius=10)
        filter_card.pack(fill="x", padx=15, pady=(5, 10))

        # ── Row 1: Date Range ──
        row1 = ctk.CTkFrame(filter_card, fg_color="transparent")
        row1.pack(fill="x", padx=16, pady=(14, 8))

        ctk.CTkLabel(row1, text="From:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 6))
        self.inp_start_date = DateEntry(
            row1, width=13, background="#1f293d", foreground="white",
            borderwidth=2, date_pattern="yyyy-mm-dd", font=("Arial", 11),
            calendar_cursor="hand2", state="readonly"
        )
        self.inp_start_date.set_date(month_start)
        self.inp_start_date.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(row1, text="To:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 6))
        self.inp_end_date = DateEntry(
            row1, width=13, background="#1f293d", foreground="white",
            borderwidth=2, date_pattern="yyyy-mm-dd", font=("Arial", 11),
            calendar_cursor="hand2", state="readonly"
        )
        self.inp_end_date.set_date(today)
        self.inp_end_date.pack(side="left", padx=(0, 14))

        ctk.CTkButton(
            row1, text="Generate Report", height=32, width=130, corner_radius=6,
            fg_color=COL_ACCENT, hover_color="#2563eb",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.generate_report
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            row1, text="Download Excel", height=32, width=130, corner_radius=6,
            fg_color="#1f293d", hover_color="#2d3d5a",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.export_excel
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            row1, text="Download PDF", height=32, width=130, corner_radius=6,
            fg_color="#3a1c1c", hover_color="#5a2d2d", text_color="#ff9f9f",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.export_pdf
        ).pack(side="left")

        # ── Row 2: Quick Presets ──
        row2 = ctk.CTkFrame(filter_card, fg_color="transparent")
        row2.pack(fill="x", padx=16, pady=(0, 8))

        ctk.CTkLabel(row2, text="Quick Range:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 10))

        presets = get_period_presets()
        for label, (start, end) in presets.items():
            ctk.CTkButton(
                row2, text=label, height=30, width=110, corner_radius=6,
                fg_color=COL_BG_INPUT, hover_color="#23283a",
                font=ctk.CTkFont(size=11),
                command=lambda s=start, e=end: self.apply_preset(s, e)
            ).pack(side="left", padx=(0, 8))

        # ── Row 3: Year Selector ──
        row3 = ctk.CTkFrame(filter_card, fg_color="transparent")
        row3.pack(fill="x", padx=16, pady=(0, 14))

        ctk.CTkLabel(row3, text="Full Year:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 10))

        years = get_available_years()
        self.year_var = ctk.StringVar(value=str(years[0]) if years else str(today.year))
        self.year_dropdown = ctk.CTkOptionMenu(
            row3, values=[str(y) for y in years], variable=self.year_var,
            width=100, height=30, fg_color=COL_BG_INPUT, button_color="#23283a",
            button_hover_color="#2d3d5a", font=ctk.CTkFont(size=12)
        )
        self.year_dropdown.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            row3, text="View Full Year", height=30, width=130, corner_radius=6,
            fg_color=COL_BG_INPUT, hover_color="#23283a",
            font=ctk.CTkFont(size=11),
            command=self.apply_year_selection
        ).pack(side="left")

        # ── Status label ──
        self.status_lbl = ctk.CTkLabel(filter_card, text="",
                                        font=ctk.CTkFont(size=12), text_color=COL_TEXT_MUTED)
        self.status_lbl.pack(anchor="w", padx=16, pady=(0, 10))

        # =================================================================
        # RESULTS TABLE (scrollable)
        # =================================================================
        table_card = ctk.CTkFrame(self, fg_color=COL_BG_CARD,
                                   border_color=COL_BORDER, border_width=1, corner_radius=10)
        table_card.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        # Header row
        header_row = ctk.CTkFrame(table_card, fg_color=COL_BG_CARD_ALT, corner_radius=0)
        header_row.pack(fill="x")

        headers = ["Date", "Product", "Available", "Sold", "Remaining", "Revenue", "Profit"]
        weights = [2, 3, 2, 2, 2, 2, 2]
        for i, w in enumerate(weights):
            header_row.grid_columnconfigure(i, weight=w)
        for col, text in enumerate(headers):
            ctk.CTkLabel(header_row, text=text, font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=COL_TEXT_MAIN).grid(row=0, column=col, sticky="w", padx=10, pady=10)

        # Scrollable body
        self.scroll_body = ctk.CTkScrollableFrame(table_card, fg_color="transparent")
        self.scroll_body.pack(fill="both", expand=True)
        for i, w in enumerate(weights):
            self.scroll_body.grid_columnconfigure(i, weight=w)

        self.col_weights = weights

        # ── Pagination bar ──
        pager_row = ctk.CTkFrame(self, fg_color="transparent")
        pager_row.pack(fill="x", padx=15, pady=(0, 10))

        self.btn_prev_page = ctk.CTkButton(
            pager_row, text="◀ Previous", width=110, height=32, corner_radius=6,
            fg_color=COL_BG_INPUT, hover_color="#23283a",
            font=ctk.CTkFont(size=12), command=self.prev_page
        )
        self.btn_prev_page.pack(side="left", padx=(0, 10))

        self.page_lbl = ctk.CTkLabel(pager_row, text="Page 1 of 1",
                                      font=ctk.CTkFont(size=12, weight="bold"),
                                      text_color=COL_TEXT_MAIN)
        self.page_lbl.pack(side="left", padx=10)

        self.btn_next_page = ctk.CTkButton(
            pager_row, text="Next ▶", width=110, height=32, corner_radius=6,
            fg_color=COL_BG_INPUT, hover_color="#23283a",
            font=ctk.CTkFont(size=12), command=self.next_page
        )
        self.btn_next_page.pack(side="left", padx=(0, 10))

        # Initial load
        self.generate_report()

    # -----------------------------------------------------------------
    # Filter actions
    # -----------------------------------------------------------------
    def apply_preset(self, start, end):
        self.inp_start_date.set_date(datetime.datetime.strptime(start, "%Y-%m-%d"))
        self.inp_end_date.set_date(datetime.datetime.strptime(end, "%Y-%m-%d"))
        self.generate_report()

    def apply_year_selection(self):
        year = int(self.year_var.get())
        start, end = get_year_range(year)
        start_dt = datetime.datetime.strptime(start, "%Y-%m-%d")
        end_dt = datetime.datetime.strptime(end, "%Y-%m-%d")
        # Don't go beyond today for the current year
        today = datetime.date.today()
        if end_dt.date() > today:
            end_dt = datetime.datetime.combine(today, datetime.time())
        self.inp_start_date.set_date(start_dt)
        self.inp_end_date.set_date(end_dt)
        self.generate_report()

    # -----------------------------------------------------------------
    # Report generation
    # -----------------------------------------------------------------
    def generate_report(self):
        start = self.inp_start_date.get_date().strftime("%Y-%m-%d")
        end = self.inp_end_date.get_date().strftime("%Y-%m-%d")

        if start > end:
            messagebox.showwarning("Invalid Range", "'From' date must be before or equal to 'To' date.")
            return

        try:
            self.cached_rows = get_product_daily_report(start, end)
        except Exception as e:
            messagebox.showerror("Report Error", f"Could not generate report:\n{e}")
            return

        self.current_page = 0
        self.render_table()

    def render_table(self):
        for widget in self.scroll_body.winfo_children():
            widget.destroy()

        if not self.cached_rows:
            self.status_lbl.configure(text="No data available for the selected range.")
            self.page_lbl.configure(text="Page 0 of 0")
            self.btn_prev_page.configure(state="disabled")
            self.btn_next_page.configure(state="disabled")
            return

        total_rows = len(self.cached_rows)
        total_pages = max(1, (total_rows + self.page_size - 1) // self.page_size)

        # Clamp current page within valid range
        if self.current_page < 0:
            self.current_page = 0
        if self.current_page > total_pages - 1:
            self.current_page = total_pages - 1

        start_idx = self.current_page * self.page_size
        end_idx = min(start_idx + self.page_size, total_rows)
        page_rows = self.cached_rows[start_idx:end_idx]

        self.status_lbl.configure(
            text=f"{total_rows} rows total | Range: {self.cached_rows[0]['date']} to {self.cached_rows[-1]['date']}",
            text_color=COL_TEXT_MUTED
        )
        self.page_lbl.configure(text=f"Page {self.current_page + 1} of {total_pages}")
        self.btn_prev_page.configure(state="normal" if self.current_page > 0 else "disabled")
        self.btn_next_page.configure(state="normal" if self.current_page < total_pages - 1 else "disabled")

        for r_idx, row in enumerate(page_rows):
            bg = COL_BG_CARD if r_idx % 2 == 0 else COL_BG_CARD_ALT

            values = [
                row["date"],
                row["product_name"],
                str(row["available"]),
                str(row["sold"]),
                str(row["remaining"]),
                f"Rs {row['revenue']:,.2f}",
                f"Rs {row['profit']:,.2f}",
            ]

            profit_color = COL_SUCCESS if row["profit"] >= 0 else COL_DANGER

            for c_idx, val in enumerate(values):
                color = COL_TEXT_MAIN
                if c_idx == 6:  # profit column
                    color = profit_color

                cell = ctk.CTkLabel(self.scroll_body, text=val, font=ctk.CTkFont(size=12),
                                     text_color=color, anchor="w", fg_color=bg)
                cell.grid(row=r_idx, column=c_idx, sticky="ew", padx=10, pady=6)

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.render_table()

    def next_page(self):
        total_pages = max(1, (len(self.cached_rows) + self.page_size - 1) // self.page_size)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self.render_table()

    # -----------------------------------------------------------------
    # Excel export
    # -----------------------------------------------------------------
    def export_excel(self):
        if not XLSX_AVAILABLE:
            messagebox.showerror("Export Failed", "openpyxl is not installed.")
            return

        if not self.cached_rows:
            messagebox.showwarning("No Data", "Generate a report first before exporting.")
            return

        start = self.inp_start_date.get_date().strftime("%Y-%m-%d")
        end = self.inp_end_date.get_date().strftime("%Y-%m-%d")
        default_name = f"product_report_{start}_to_{end}.xlsx"

        save_path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Product Report As",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=(("Excel Workbook", "*.xlsx"), ("All Files", "*.*"))
        )

        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Product Report"

            headers = ["Date", "Product", "Available", "Sold", "Remaining", "Revenue", "Cost", "Profit"]
            ws.append(headers)

            header_fill = PatternFill(start_color="1F293D", end_color="1F293D", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in self.cached_rows:
                ws.append([
                    row["date"], row["product_name"], row["available"], row["sold"],
                    row["remaining"], row["revenue"], row["cost"], row["profit"]
                ])

            for col_idx, header in enumerate(headers, start=1):
                max_len = len(str(header))
                for row in self.cached_rows:
                    val = list(row.values())[col_idx - 1] if col_idx - 1 < len(row) else ""
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

            wb.save(save_path)
            messagebox.showinfo("Export Successful", f"Report saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save report:\n{e}")

    # -----------------------------------------------------------------
    # PDF export
    # -----------------------------------------------------------------
    def export_pdf(self):
        if not self.cached_rows:
            messagebox.showwarning("No Data", "Generate a report first before exporting.")
            return

        start = self.inp_start_date.get_date().strftime("%Y-%m-%d")
        end = self.inp_end_date.get_date().strftime("%Y-%m-%d")
        default_name = f"product_report_{start}_to_{end}.pdf"

        save_path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Product Report As PDF",
            initialfile=default_name,
            defaultextension=".pdf",
            filetypes=(("PDF File", "*.pdf"), ("All Files", "*.*"))
        )

        if not save_path:
            return

        try:
            headers = ["Date", "Product", "Available", "Sold", "Remaining", "Revenue", "Cost", "Profit"]
            col_classes = ["", "", "right", "right", "right", "right", "right", "right"]
            col_widths = [11, 23, 10, 9, 11, 13, 11, 12]

            rows = []
            total_revenue = 0
            total_cost = 0
            total_profit = 0

            for row in self.cached_rows:
                total_revenue += row["revenue"]
                total_cost += row["cost"]
                total_profit += row["profit"]

                rows.append([
                    row["date"],
                    row["product_name"],
                    str(row["available"]),
                    str(row["sold"]),
                    str(row["remaining"]),
                    f"{row['revenue']:,.2f}",
                    f"{row['cost']:,.2f}",
                    (f"{row['profit']:,.2f}", "green" if row["profit"] >= 0 else "red"),
                ])

            totals_row = [
                ("TOTALS", "bold"), "", "", "", "",
                (f"{total_revenue:,.2f}", "green"),
                (f"{total_cost:,.2f}", ""),
                (f"{total_profit:,.2f}", "green" if total_profit >= 0 else "red"),
            ]

            summary_rows = [
                ("Report Range:", f"{start} to {end}"),
                ("Total Rows:", str(len(self.cached_rows))),
                ("Total Revenue:", f"Rs. {total_revenue:,.2f}"),
                ("Total Cost:", f"Rs. {total_cost:,.2f}"),
                ("Total Profit:", f"Rs. {total_profit:,.2f}"),
            ]

            html = build_table_html(
                title="📦 Product Daily Report",
                subtitle=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                headers=headers,
                rows=rows,
                col_classes=col_classes,
                orientation="landscape",
                summary_rows=summary_rows,
                totals_row=totals_row,
                col_widths=col_widths,
            )

            success, msg = export_html_to_pdf(html, save_path)
            if success:
                messagebox.showinfo("Export Successful", f"Report saved to:\n{save_path}")
                try:
                    import os
                    os.startfile(save_path)
                except Exception:
                    pass
            else:
                messagebox.showerror("Export Failed", msg)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save PDF:\n{e}")
