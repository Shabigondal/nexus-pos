import customtkinter as ctk
import database.db_manager as db
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import os

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

class InventoryView(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")
        
        # 🔢 INVENTORY PAGINATION UTILITY MATRIX STATES
        self.current_page = 0
        self.items_per_page = 10
        self.cached_products = [] # Active text query runtime dataset storage array
        
        # --- HEADER AND ADD ACTION CONTROL ---
        self.top_bar = ctk.CTkFrame(self, fg_color="transparent")
        self.top_bar.pack(fill="x", padx=20, pady=(10, 15))
        
        self.search_entry = ctk.CTkEntry(self.top_bar, placeholder_text="🔍 Search products by title or barcode context...", 
                                         height=40, fg_color="#16161a", border_color="#222227", corner_radius=6)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(0, 15))
        self.search_entry.bind("<KeyRelease>", self.trigger_live_search)
        
        self.btn_add_pop = ctk.CTkButton(self.top_bar, text="+ Add New Product", height=40, width=160,
                                         fg_color="#1f293d", hover_color="#2d3d5a", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                                         corner_radius=6, command=self.open_product_modal)
        self.btn_add_pop.pack(side="right", padx=(6, 0))

        ctk.CTkButton(self.top_bar, text="📥 Import Excel", height=40, width=130,
                      fg_color="#1c3a27", hover_color="#1f4a2f", text_color="#4aff4a",
                      border_color="#2d5a3b", border_width=1,
                      font=ctk.CTkFont(size=12, weight="bold"),
                      corner_radius=6, command=self.import_from_excel).pack(side="right", padx=6)

        ctk.CTkButton(self.top_bar, text="📤 Export Excel", height=40, width=130,
                      fg_color="#1f293d", hover_color="#2d3d5a", text_color="#8ed1fc",
                      border_color="#2a3a5a", border_width=1,
                      font=ctk.CTkFont(size=12, weight="bold"),
                      corner_radius=6, command=self.export_to_excel).pack(side="right", padx=(0, 0))
        
        # --- CENTRAL UNIFIED DATA GRID ---
        self.grid_container = ctk.CTkScrollableFrame(self, fg_color="#121214", border_color="#222227", border_width=1, corner_radius=6)
        self.grid_container.pack(fill="both", expand=True, padx=20, pady=(0, 5))
        
        # 🎛️ ORIGINAL THEME COMPATIBLE PAGINATION FOOTER ROW BLOCK
        self.paginator_frame = ctk.CTkFrame(self, fg_color="transparent", height=40)
        self.paginator_frame.pack(fill="x", padx=20, pady=(10, 15))
        
        self.btn_prev = ctk.CTkButton(self.paginator_frame, text="⬅️ Previous Page", width=120, height=32, 
                                      fg_color="#16161a", border_color="#222227", border_width=1, hover_color="#1f1f24", 
                                      font=ctk.CTkFont(size=12, weight="bold"), command=self.navigate_prev)
        self.btn_prev.pack(side="left")
        
        self.lbl_page_num = ctk.CTkLabel(self.paginator_frame, text="Page 1 of 1", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="gray")
        self.lbl_page_num.pack(side="left", expand=True)
        
        self.btn_next = ctk.CTkButton(self.paginator_frame, text="Next Page ➡️", width=120, height=32, 
                                      fg_color="#16161a", border_color="#222227", border_width=1, hover_color="#1f1f24", 
                                      font=ctk.CTkFont(size=12, weight="bold"), command=self.navigate_next)
        self.btn_next.pack(side="right")
        
        # Core Initialization Data Execution Pull
        self.refresh_data_grid()

    def refresh_data_grid(self, search_text=""):
        # Fetch unmatched dataset records stream directly into internal frame cache memory objects
        self.cached_products = db.get_all_products(search_text)
        self.render_paginated_rows_matrix()

    def render_paginated_rows_matrix(self):
        # Clear existing dynamic grid node components rows
        for widget in self.grid_container.winfo_children():
            widget.destroy()
            
        total_items = len(self.cached_products)
        max_pages = max(1, (total_items + self.items_per_page - 1) // self.items_per_page)
        
        # Safety normalization boundary controls check
        if self.current_page >= max_pages:
            self.current_page = max_pages - 1
            
        # Paginator layout states sync processing
        self.lbl_page_num.configure(text=f"Page {self.current_page + 1} of {max_pages}")
        self.btn_prev.configure(state="normal" if self.current_page > 0 else "disabled")
        self.btn_next.configure(state="normal" if (self.current_page + 1) < max_pages else "disabled")
        
        # Extract exact 10 records range tracking slice indexes
        start_idx = self.current_page * self.items_per_page
        end_idx = start_idx + self.items_per_page
        page_chunk = self.cached_products[start_idx:end_idx]
        
        # Definitive Enterprise Column Width Matrix (Keep your exact dimensions safe)
        widths = [35, 95, 190, 55, 55, 80, 80, 85, 70]
        
        # Table Header Row Build
        header_frame = ctk.CTkFrame(self.grid_container, fg_color="#16161a", corner_radius=4, height=35)
        header_frame.pack(fill="x", pady=(0, 8))
        
        headers = ["S.No", "Barcode", "Product Title", "Unit", "Stock", "Cost Price", "Sale Price", "Status", "Management Actions"]
        for idx, text in enumerate(headers):
            anchor_pos = "w" if idx == 2 else "center"
            lbl = ctk.CTkLabel(header_frame, text=text, width=widths[idx], font=ctk.CTkFont(size=12, weight="bold"), text_color="#4a90e2", anchor=anchor_pos)
            lbl.pack(side="left", padx=4)

        # Populate logical records chunk elements (Exact your original UI row style matching schema)
        for idx, p in enumerate(page_chunk):
            p_id, name, barcode, unit, qty, cost, sale = p
            display_barcode = barcode if barcode else "N/A"
            
            # Global sequence serial tracker accounting calculation based on dynamic math bounds
            global_serial_no = start_idx + idx + 1
            
            row_frame = ctk.CTkFrame(self.grid_container, fg_color="#18181f" if idx % 2 == 0 else "transparent", height=45, corner_radius=4)
            row_frame.pack(fill="x", pady=2)
            
            # Text Fields Generation
            ctk.CTkLabel(row_frame, text=str(global_serial_no), width=widths[0], text_color="gray").pack(side="left", padx=4)
            ctk.CTkLabel(row_frame, text=display_barcode, width=widths[1], text_color="#a0a0a9").pack(side="left", padx=4)
            ctk.CTkLabel(row_frame, text=name, width=widths[2], anchor="w", text_color="#ffffff", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=4)
            ctk.CTkLabel(row_frame, text=unit, width=widths[3], text_color="#8ed1fc").pack(side="left", padx=4)
            qty_display = str(int(qty)) if float(qty) == int(qty) else f"{round(float(qty), 3):g}"
            ctk.CTkLabel(row_frame, text=qty_display, width=widths[4], text_color="#ffffff").pack(side="left", padx=4)
            ctk.CTkLabel(row_frame, text=f"Rs {cost:,.2f}", width=widths[5], text_color="#f0ad4e").pack(side="left", padx=4)
            ctk.CTkLabel(row_frame, text=f"Rs {sale:,.2f}", width=widths[6], text_color="#4aff8e").pack(side="left", padx=4)
            
            # Dynamic Inventory Shortage Alert System Engine
            status_frame = ctk.CTkFrame(row_frame, width=widths[7], height=26, corner_radius=4)
            status_frame.pack(side="left", padx=4)
            status_frame.pack_propagate(False)
            
            if qty <= 5:
                status_frame.configure(fg_color="#3a1c1c")
                status_text = "SHORTAGE"
                status_color = "#ff4a4a"
            else:
                status_frame.configure(fg_color="#1c3a27")
                status_text = "OPTIMAL"
                status_color = "#4aff4a"
                
            ctk.CTkLabel(status_frame, text=status_text, text_color=status_color, font=ctk.CTkFont(size=10, weight="bold")).pack(expand=True)
            
            # Management Action Interface Buttons (Edit & Delete Inline, icon-only)
            actions_frame = ctk.CTkFrame(row_frame, fg_color="transparent", width=widths[8])
            actions_frame.pack(side="right", padx=4)
            
            btn_edit = ctk.CTkButton(actions_frame, text="✏️", width=30, height=26, fg_color="#2b3a4a", hover_color="#3d5268", corner_radius=4,
                                      font=ctk.CTkFont(size=13),
                                      command=lambda prod=p: self.open_product_modal(prod))
            btn_edit.pack(side="left", padx=2)
            
            btn_del = ctk.CTkButton(actions_frame, text="🗑️", width=30, height=26, fg_color="#3a1c1c", hover_color="#542424", text_color="#ff8080", corner_radius=4,
                                     font=ctk.CTkFont(size=13),
                                     command=lambda target_id=p_id: self.trigger_purge(target_id))
            btn_del.pack(side="left", padx=2)

    def trigger_live_search(self, event):
        self.current_page = 0 # Reset frame cursor indexing back to page zero context
        self.refresh_data_grid(self.search_entry.get().strip())

    def trigger_purge(self):
        # Fallback security checker over targeted id blocks maps executions
        pass

    def trigger_purge(self, target_id):
        db.delete_product(target_id)
        self.refresh_data_grid(self.search_entry.get().strip())

    def navigate_prev(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.render_paginated_rows_matrix()

    def navigate_next(self):
        total_items = len(self.cached_products)
        max_pages = (total_items + self.items_per_page - 1) // self.items_per_page
        if (self.current_page + 1) < max_pages:
            self.current_page += 1
            self.render_paginated_rows_matrix()

    # =================================================================
    # 📑 MODAL DIALOG ENVIRONMENT (ADD / EDIT FORM POPUP)
    # =================================================================
    def open_product_modal(self, edit_payload=None):
        modal = ctk.CTkToplevel(self)
        modal.title("Asset Configuration Wizard" if edit_payload else "Register Product Asset")
        modal.geometry("450x580")
        modal.resizable(False, False)
        modal.attributes("-topmost", True)
        modal.grab_set()
        modal.configure(fg_color="#121214")
        
        wrapper = ctk.CTkFrame(modal, fg_color="#1a1a1e", border_color="#222227", border_width=1, corner_radius=6)
        wrapper.pack(fill="both", expand=True, padx=20, pady=20)
        
        title_txt = "Modify Product Context" if edit_payload else "New Product Initialization"
        lbl_title = ctk.CTkLabel(wrapper, text=title_txt, font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"), text_color="#ffffff")
        lbl_title.pack(pady=(25, 20))
        
        # Fields Deployment Matrix
        inp_name = self.build_modal_input(wrapper, "Product Description / Name *")
        inp_barcode = self.build_modal_input(wrapper, "System Barcode Identifier (Optional)")
        
        lbl_u = ctk.CTkLabel(wrapper, text="Measurement Metric Unit *", font=ctk.CTkFont(size=11), text_color="gray")
        lbl_u.pack(anchor="w", padx=40, pady=(5, 2))
        inp_unit = ctk.CTkOptionMenu(wrapper, values=["KG", "Litre", "Pieces", "Pack", "Carton"], width=340, height=38, fg_color="#121214", button_color="#222227", corner_radius=4)
        inp_unit.pack(pady=(0, 8))
        
        inp_qty = self.build_modal_input(wrapper, "Current Stock Quantity Matrix *")
        inp_cost = self.build_modal_input(wrapper, "Unit Acquisition Base Cost Price *")
        inp_sale = self.build_modal_input(wrapper, "Target Market Retail Sale Price *")
        
        lbl_err = ctk.CTkLabel(wrapper, text="", font=ctk.CTkFont(size=11))
        lbl_err.pack(pady=2)

        # Pre-load features if Edit Payload exists
        if edit_payload:
            inp_name.insert(0, edit_payload[1])
            if edit_payload[2]: inp_barcode.insert(0, edit_payload[2])
            inp_unit.set(edit_payload[3])
            inp_qty.insert(0, str(edit_payload[4]))
            inp_cost.insert(0, str(edit_payload[5]))
            inp_sale.insert(0, str(edit_payload[6]))

        def save_transaction():
            n = inp_name.get().strip()
            b = inp_barcode.get().strip()
            u = inp_unit.get()
            q = inp_qty.get().strip()
            c = inp_cost.get().strip()
            s = inp_sale.get().strip()
            
            # Compulsory Field Checks
            if not n or not q or not c or not s:
                lbl_err.configure(text="Execution Error: Mandatory fields required.", text_color="#ff4a4a")
                return
                
            try:
                # Type validation mechanics
                float(q)
                float(c)
                float(s)
            except ValueError:
                lbl_err.configure(text="Validation Mismatch: Quantities/Prices must be numeric.", text_color="#ff4a4a")
                return
                
            if edit_payload:
                success, msg = db.update_product(edit_payload[0], n, b, u, q, c, s)
            else:
                success, msg = db.add_product(n, b, u, q, c, s)
                
            if success:
                modal.destroy()
                self.refresh_data_grid(self.search_entry.get().strip())
            else:
                lbl_err.configure(text=msg, text_color="#ff4a4a")

        btn_action_txt = "Apply Core Mutations" if edit_payload else "Commit Asset to Database"
        btn_save = ctk.CTkButton(wrapper, text=btn_action_txt, height=42, fg_color="#1f293d", hover_color="#2d3d5a", font=ctk.CTkFont(weight="bold"), corner_radius=4, command=save_transaction)
        btn_save.pack(fill="x", padx=40, pady=(15, 5))

    def build_modal_input(self, parent, placeholder):
        entry = ctk.CTkEntry(parent, placeholder_text=placeholder, width=340, height=38, fg_color="#121214", border_color="#222227", corner_radius=4)
        entry.pack(pady=6)
        return entry

    # =================================================================
    # 📤 EXPORT INVENTORY TO EXCEL
    # =================================================================
    def export_to_excel(self):
        if not XLSX_AVAILABLE:
            messagebox.showerror("Missing Library", "openpyxl install nahi hai.\npip install openpyxl")
            return

        all_products = db.get_all_products()
        if not all_products:
            messagebox.showwarning("No Data", "Inventory mein koi product nahi hai.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")],
            initialfile="Inventory_Export.xlsx",
            title="Export Inventory"
        )
        if not save_path: return

        try:
            import datetime
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"
            ws.sheet_view.showGridLines = False

            hdr_fill  = PatternFill("solid", fgColor="1e3a5f")
            hdr_font  = Font(color="FFFFFF", bold=True, size=11, name="Arial")
            body_font = Font(color="f1f5f9", name="Arial", size=10)
            muted     = Font(color="94a3b8", name="Arial", size=10, italic=True)
            center    = Alignment(horizontal="center", vertical="center")
            right_aln = Alignment(horizontal="right")
            border    = Border(
                left=Side(style="thin", color="2a2e38"),
                right=Side(style="thin", color="2a2e38"),
                top=Side(style="thin", color="2a2e38"),
                bottom=Side(style="thin", color="2a2e38")
            )

            # Title
            ws.merge_cells("A1:H1")
            c = ws["A1"]
            c.value = f"📦 Inventory Export  —  {datetime.date.today()}"
            c.font = Font(color="FFFFFF", bold=True, size=14, name="Arial")
            c.fill = PatternFill("solid", fgColor="0f1117")
            c.alignment = center
            ws.row_dimensions[1].height = 28

            # Instruction row
            ws.merge_cells("A2:H2")
            c = ws["A2"]
            c.value = "ℹ️  Import ke liye: S.No column mat badlo | Product Name exact rakho | Unit: KG / Litre / Pieces / Pack / Carton"
            c.font = muted
            c.fill = PatternFill("solid", fgColor="11131a")
            c.alignment = center
            ws.row_dimensions[2].height = 18

            # Headers — same order as import template
            hdrs = ["S.No", "Product Name *", "Barcode", "Unit *", "Quantity *", "Cost Price *", "Sale Price *", "Status"]
            ws.row_dimensions[3].height = 22
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = center; c.border = border

            for ri, p in enumerate(all_products, 4):
                p_id, name, barcode, unit, qty, cost, sale = p
                fill = PatternFill("solid", fgColor="1a1d24") if ri % 2 == 0 else PatternFill("solid", fgColor="15171d")
                status = "SHORTAGE" if qty <= 5 else "OPTIMAL"
                status_color = "ff4a4a" if qty <= 5 else "4aff4a"

                row_vals = [ri - 3, name, barcode or "", unit, round(qty, 4), round(cost, 2), round(sale, 2), status]
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.fill = fill; c.border = border
                    if ci == 8:
                        c.font = Font(color=status_color, bold=True, name="Arial", size=10)
                        c.alignment = center
                    elif ci in (6, 7):
                        c.font = body_font; c.number_format = '#,##0.00'; c.alignment = right_aln
                    elif ci in (1, 4, 5):
                        c.font = body_font; c.alignment = center
                    else:
                        c.font = body_font

            col_widths = [6, 32, 16, 10, 10, 14, 14, 12]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(save_path)
            messagebox.showinfo("Export Successful ✅",
                f"{len(all_products)} products export ho gaye!\n\nFile:\n{save_path}\n\n"
                "Tip: Yahi file edit karke Import Excel se wapas upload kar saktay hain.")
            try: os.startfile(save_path)
            except Exception: pass

        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    # =================================================================
    # 📥 IMPORT INVENTORY FROM EXCEL — PROFESSIONAL UPSERT ENGINE
    # =================================================================
    def import_from_excel(self):
        if not XLSX_AVAILABLE and not PANDAS_AVAILABLE:
            messagebox.showerror("Missing Library",
                "openpyxl ya pandas install nahi hain.\npip install openpyxl pandas")
            return

        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
            title="Import Inventory from Excel"
        )
        if not file_path: return

        try:
            # Read using openpyxl directly — no pandas dependency needed
            if PANDAS_AVAILABLE:
                df = pd.read_excel(file_path, header=2)
            elif XLSX_AVAILABLE:
                wb_in = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws_in = wb_in.active
                all_rows = list(ws_in.iter_rows(min_row=3, values_only=True))
                if len(all_rows) < 2:
                    messagebox.showwarning("No Data", "File mein koi product row nahi mili.")
                    return
                col_names = [str(c).strip().lower().replace(" *", "").replace(" ", "_") if c else "" for c in all_rows[0]]
                data_rows = []
                for r in all_rows[1:]:
                    data_rows.append(dict(zip(col_names, r)))
                # Convert to simple list-of-dicts for processing below
                import types
                df = types.SimpleNamespace()
                df._rows = data_rows
                df.columns = col_names
                df._is_simple = True
            else:
                messagebox.showerror("Missing Library", "openpyxl ya pandas install nahi hain.\npip install openpyxl")
                return

            # Normalize column names
            if not getattr(df, '_is_simple', False):
                df.columns = [str(c).strip().lower().replace(" *", "").replace(" ", "_") for c in df.columns]

            # Expected columns check
            required = {"product_name", "unit", "quantity", "cost_price", "sale_price"}
            col_set = set(df.columns)
            missing = required - col_set
            if missing:
                messagebox.showerror("Column Error",
                    f"Ye columns nahi mile file mein:\n{', '.join(missing)}\n\n"
                    "Pehle Export Excel se template download karein aur usi format mein fill karein.")
                return

            # Build unified row iterator
            if getattr(df, '_is_simple', False):
                row_iter = enumerate(df._rows)
            else:
                df = df.dropna(subset=["product_name"])
                df = df[df["product_name"].astype(str).str.strip() != ""]
                df = df[df["product_name"].astype(str).str.strip() != "nan"]
                row_iter = df.iterrows()

            if getattr(df, '_is_simple', False) and not df._rows:
                messagebox.showwarning("No Data", "File mein koi valid product row nahi mili.")
                return

            valid_units = {"KG", "Litre", "Pieces", "Pack", "Carton"}

            added = 0
            updated = 0
            skipped = 0
            errors = []

            # Fetch existing products for name-based duplicate check
            existing = db.get_all_products()
            existing_map = {p[1].strip().lower(): p for p in existing}  # name → full tuple

            for row_num, row in row_iter:
                try:
                    # Support both dict (SimpleNamespace) and pandas Series
                    def get_val(r, key, default=""):
                        if isinstance(r, dict):
                            return r.get(key, default)
                        return r.get(key, default) if hasattr(r, 'get') else getattr(r, key, default)

                    name     = str(get_val(row, "product_name", "")).strip()
                    barcode  = str(get_val(row, "barcode", "")).strip()
                    unit     = str(get_val(row, "unit", "Pieces")).strip()
                    qty_raw  = get_val(row, "quantity", 0)
                    cost_raw = get_val(row, "cost_price", 0)
                    sale_raw = get_val(row, "sale_price", 0)

                    if not name or name.lower() == "nan":
                        continue

                    # Barcode cleanup
                    if barcode.lower() in ("nan", "none", ""):
                        barcode = ""

                    # Unit validation — default to Pieces if invalid
                    if unit not in valid_units:
                        unit = "Pieces"

                    # Numeric validation
                    try:
                        qty  = round(float(str(qty_raw)), 4)
                        cost = float(str(cost_raw))
                        sale = float(str(sale_raw))
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num + 1}: '{name}' — price/qty numeric nahi hai, skip.")
                        skipped += 1
                        continue

                    if qty < 0 or cost < 0 or sale < 0:
                        errors.append(f"Row {row_num + 1}: '{name}' — negative values nahi chalenge, skip.")
                        skipped += 1
                        continue

                    name_key = name.lower()

                    if name_key in existing_map:
                        # ── UPSERT: product already exists → quantity ADD karo, prices update karo
                        existing_prod = existing_map[name_key]
                        p_id         = existing_prod[0]
                        old_qty      = existing_prod[4]
                        new_qty      = old_qty + qty  # stock add hoga, replace nahi
                        # barcode: agar file mein diya hai toh update, warna purana rakho
                        new_barcode  = barcode if barcode else (existing_prod[2] or "")
                        db.update_product(p_id, name, new_barcode, unit, new_qty, cost, sale)
                        updated += 1
                        # Update local map so duplicate rows in same file stack correctly
                        existing_map[name_key] = (p_id, name, new_barcode, unit, new_qty, cost, sale)
                    else:
                        # ── NEW product
                        ok, msg = db.add_product(name, barcode, unit, qty, cost, sale)
                        if ok:
                            added += 1
                            # Add to map for intra-file duplicate handling
                            existing_map[name_key] = (None, name, barcode, unit, qty, cost, sale)
                        else:
                            errors.append(f"Row {row_num + 1}: '{name}' — {msg}")
                            skipped += 1

                except Exception as row_err:
                    errors.append(f"Row {row_num + 1}: unexpected error — {row_err}")
                    skipped += 1

            # Refresh grid
            self.current_page = 0
            self.refresh_data_grid()

            # Result summary
            summary = (
                f"✅ Import Complete!\n\n"
                f"• Naye products add: {added}\n"
                f"• Existing stock updated (qty add): {updated}\n"
                f"• Skipped (errors): {skipped}"
            )
            if errors:
                summary += f"\n\n⚠️ Issues ({len(errors)}):\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    summary += f"\n...aur {len(errors) - 5} aur"

            messagebox.showinfo("Import Result", summary)

        except Exception as e:
            messagebox.showerror("Import Failed",
                f"File read nahi ho saki:\n{str(e)}\n\n"
                "Make sure file Export Excel se generate ki gayi ho ya same format mein ho.")