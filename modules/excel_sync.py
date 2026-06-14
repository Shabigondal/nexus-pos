"""
Excel Auto-Sync Module
------------------------
Keeps two Excel files automatically up to date with the live database:

  - products.xlsx  -> mirrors the `inventory` table (billing_system.db)
  - khata.xlsx     -> mirrors the `ledger_customers` table (database/pos_system.db)

Both files are rewritten in full every time a relevant add/update/delete
happens, so they always reflect the current state of the database.

The files are saved in the application's root folder by default.
"""

import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database.db_manager import DB_PATH, get_all_products, get_all_ledger_customers_full

PRODUCTS_XLSX = "products.xlsx"
KHATA_XLSX = "khata.xlsx"

LEDGER_DB_PATH = os.path.join("database", "pos_system.db")

HEADER_FILL = PatternFill(start_color="1F293D", end_color="1F293D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# --------------------------------------------------------------------------
# Internal helpers
# --------------------------------------------------------------------------
def _write_sheet(path, headers, rows, sheet_title="Sheet1"):
    """Creates/overwrites an .xlsx file with a single sheet of headers + rows."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append(list(row))

        # Auto-fit column widths (simple heuristic)
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4

        wb.save(path)
        return True, None
    except Exception as e:
        return False, str(e)


# --------------------------------------------------------------------------
# PRODUCTS SYNC
# --------------------------------------------------------------------------
def sync_products_excel():
    """
    Rewrites products.xlsx with the current contents of the inventory table.
    Columns: Product ID, Name, Barcode, Unit, Quantity, Cost Price, Sale Price
    """
    try:
        rows = get_all_products()
        headers = ["Product ID", "Name", "Barcode", "Unit", "Quantity", "Cost Price", "Sale Price"]
        return _write_sheet(PRODUCTS_XLSX, headers, rows, sheet_title="Products")
    except Exception as e:
        return False, str(e)


# --------------------------------------------------------------------------
# KHATA SYNC
# --------------------------------------------------------------------------
def sync_khata_excel():
    """
    Rewrites khata.xlsx with the current contents of the ledger_customers table.
    Columns: Khata ID, Customer Name, Phone Number, Wallet Balance, Created At
    """
    try:
        rows = get_all_ledger_customers_full()
        headers = ["Khata ID", "Customer Name", "Phone Number", "Wallet Balance", "Created At"]
        return _write_sheet(KHATA_XLSX, headers, rows, sheet_title="Khata")
    except Exception as e:
        return False, str(e)


# --------------------------------------------------------------------------
# Combined helper - call after any product or khata change
# --------------------------------------------------------------------------
def sync_all_excel():
    """Re-syncs both products.xlsx and khata.xlsx. Safe to call frequently."""
    sync_products_excel()
    sync_khata_excel()
