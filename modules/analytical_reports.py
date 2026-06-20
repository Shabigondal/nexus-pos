import customtkinter as ctk
import database.db_manager as db
from modules.invoice_print import InvoicePrintWindow, export_invoice_pdf
from modules.pdf_export import build_table_html, export_html_to_pdf
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
from tkcalendar import DateEntry
import datetime
import csv
import os
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# =====================================================================
# 🎨 THEME PALETTE (matches billing.py)
# =====================================================================
COL_BG_DARK     = "#0f1117"
COL_BG_CARD     = "#1a1d24"
COL_BG_CARD_ALT = "#15171d"
COL_BG_INPUT    = "#11131a"
COL_BORDER      = "#2a2e38"
COL_ACCENT      = "#3b82f6"
COL_ACCENT_SOFT = "#1e293b"
COL_SUCCESS     = "#22c55e"
COL_SUCCESS_BG  = "#14241c"
COL_WARNING     = "#f59e0b"
COL_WARNING_BG  = "#292011"
COL_DANGER      = "#ef4444"
COL_DANGER_BG   = "#2a1518"
COL_TEXT_MAIN   = "#f1f5f9"
COL_TEXT_MUTED  = "#94a3b8"
COL_TEXT_SOFT   = "#64748b"

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

CURRENT_YEAR = datetime.date.today().year
YEARS = [str(y) for y in range(CURRENT_YEAR, CURRENT_YEAR - 6, -1)]


class AnalyticalReportsView(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color="transparent")

        self.grid_columnconfigure(0, weight=4)
        self.grid_columnconfigure(1, weight=2)
        self.grid_rowconfigure(0, weight=1)

        self.active_selection = None
        self.current_page = 0
        self.items_per_page = 12
        self.cached_invoices = []
        self.filter_mode = "date"  # "date" or "month"

        # =================================================================
        # 📊 LEFT PANEL
        # =================================================================
        self.left_pane = ctk.CTkFrame(self, fg_color="transparent")
        self.left_pane.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        # ── FILTER CARD ──────────────────────────────────────────────
        filter_card = ctk.CTkFrame(self.left_pane, fg_color=COL_BG_CARD,
                                    border_color=COL_BORDER, border_width=1, corner_radius=10)
        filter_card.pack(fill="x", padx=15, pady=(5, 10))

        # Search bar
        self.search_entry = ctk.CTkEntry(
            filter_card, placeholder_text="🔍  Search by Invoice ID or Customer Name...",
            height=40, fg_color=COL_BG_INPUT, border_color=COL_BORDER,
            corner_radius=8, font=ctk.CTkFont(size=13)
        )
        self.search_entry.pack(fill="x", padx=16, pady=(14, 8))
        self.search_entry.bind("<KeyRelease>", lambda e: self.trigger_new_search_context())

        # ── TAB ROW: Date Filter / Month Filter ──
        tab_row = ctk.CTkFrame(filter_card, fg_color="transparent")
        tab_row.pack(fill="x", padx=16, pady=(0, 8))

        self.btn_tab_date = ctk.CTkButton(
            tab_row, text="📅  Date Range", height=32, corner_radius=6,
            fg_color=COL_ACCENT, hover_color="#2563eb", text_color="#fff",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=lambda: self.switch_filter_mode("date")
        )
        self.btn_tab_date.pack(side="left", padx=(0, 6))

        self.btn_tab_month = ctk.CTkButton(
            tab_row, text="🗓️  Monthly View", height=32, corner_radius=6,
            fg_color=COL_BG_INPUT, hover_color=COL_BORDER, text_color=COL_TEXT_MUTED,
            border_color=COL_BORDER, border_width=1, font=ctk.CTkFont(size=12, weight="bold"),
            command=lambda: self.switch_filter_mode("month")
        )
        self.btn_tab_month.pack(side="left")

        # ── DATE RANGE FILTER FRAME ──
        self.date_filter_frame = ctk.CTkFrame(filter_card, fg_color=COL_BG_INPUT,
                                               border_color=COL_BORDER, border_width=1, corner_radius=8)
        self.date_filter_frame.pack(fill="x", padx=16, pady=(0, 14))

        date_inner = ctk.CTkFrame(self.date_filter_frame, fg_color="transparent")
        date_inner.pack(fill="x", padx=12, pady=10)

        ctk.CTkLabel(date_inner, text="From:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 6))
        self.inp_start_date = DateEntry(
            date_inner, width=13, background="#1f293d", foreground="white",
            borderwidth=2, date_pattern="yyyy-mm-dd", font=("Arial", 11),
            calendar_cursor="hand2", state="readonly"
        )
        self.inp_start_date.pack(side="left", padx=(0, 14))

        ctk.CTkLabel(date_inner, text="To:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 6))
        self.inp_end_date = DateEntry(
            date_inner, width=13, background="#1f293d", foreground="white",
            borderwidth=2, date_pattern="yyyy-mm-dd", font=("Arial", 11),
            calendar_cursor="hand2", state="readonly"
        )
        self.inp_end_date.pack(side="left", padx=(0, 14))

        self.btn_apply_date = ctk.CTkButton(
            date_inner, text="Apply", height=30, width=70, corner_radius=6,
            fg_color=COL_ACCENT, hover_color="#2563eb",
            font=ctk.CTkFont(size=11, weight="bold"),
            command=self.trigger_new_search_context
        )
        self.btn_apply_date.pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            date_inner, text="Reset", height=30, width=60, corner_radius=6,
            fg_color="transparent", border_color=COL_BORDER, border_width=1,
            text_color=COL_TEXT_SOFT, hover_color=COL_BG_CARD_ALT,
            font=ctk.CTkFont(size=11), command=self.clear_date_inputs
        ).pack(side="left")

        # ── MONTHLY FILTER FRAME ──
        self.month_filter_frame = ctk.CTkFrame(filter_card, fg_color=COL_BG_INPUT,
                                                border_color=COL_BORDER, border_width=1, corner_radius=8)
        # Hidden by default — shown when tab switches

        month_inner = ctk.CTkFrame(self.month_filter_frame, fg_color="transparent")
        month_inner.pack(fill="x", padx=12, pady=10)

        ctk.CTkLabel(month_inner, text="Month:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 6))

        self.month_var = ctk.StringVar(value=MONTHS[datetime.date.today().month - 1])
        self.month_menu = ctk.CTkOptionMenu(
            month_inner, variable=self.month_var, values=MONTHS,
            width=130, height=30, corner_radius=6,
            fg_color=COL_BG_CARD, button_color=COL_ACCENT, button_hover_color="#2563eb",
            font=ctk.CTkFont(size=12), command=lambda _: self.trigger_new_search_context()
        )
        self.month_menu.pack(side="left", padx=(0, 12))

        ctk.CTkLabel(month_inner, text="Year:", font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=COL_TEXT_MUTED).pack(side="left", padx=(0, 6))

        self.year_var = ctk.StringVar(value=str(CURRENT_YEAR))
        self.year_menu = ctk.CTkOptionMenu(
            month_inner, variable=self.year_var, values=YEARS,
            width=90, height=30, corner_radius=6,
            fg_color=COL_BG_CARD, button_color=COL_ACCENT, button_hover_color="#2563eb",
            font=ctk.CTkFont(size=12), command=lambda _: self.trigger_new_search_context()
        )
        self.year_menu.pack(side="left", padx=(0, 14))

        ctk.CTkButton(
            month_inner, text="Show", height=30, width=70, corner_radius=6,
            fg_color=COL_ACCENT, hover_color="#2563eb",
            font=ctk.CTkFont(size=11, weight="bold"),
            command=self.trigger_new_search_context
        ).pack(side="left")

        # ── SUMMARY STATS ROW ──
        self.stats_frame = ctk.CTkFrame(self.left_pane, fg_color="transparent")
        self.stats_frame.pack(fill="x", padx=15, pady=(0, 8))

        self.stat_count = self._make_stat_tile(self.stats_frame, "Total Invoices", "0", COL_ACCENT)
        self.stat_revenue = self._make_stat_tile(self.stats_frame, "Total Revenue", "Rs. 0", COL_SUCCESS)
        self.stat_profit = self._make_stat_tile(self.stats_frame, "Total Profit", "Rs. —", "#a855f7")
        self.stat_avg = self._make_stat_tile(self.stats_frame, "Avg. Invoice", "Rs. 0", COL_WARNING)

        # ── INVOICE TABLE ──
        self.table_container = ctk.CTkScrollableFrame(
            self.left_pane, fg_color=COL_BG_CARD,
            border_color=COL_BORDER, border_width=1, corner_radius=10
        )
        self.table_container.pack(fill="both", expand=True, padx=15, pady=(0, 4))

        # ── PAGINATION + EXPORT ROW ──
        bottom_row = ctk.CTkFrame(self.left_pane, fg_color="transparent")
        bottom_row.pack(fill="x", padx=15, pady=(4, 10))

        self.btn_prev = ctk.CTkButton(
            bottom_row, text="⬅ Prev", width=90, height=32,
            fg_color=COL_BG_CARD, border_color=COL_BORDER, border_width=1,
            text_color=COL_TEXT_MUTED, hover_color=COL_BG_CARD_ALT,
            font=ctk.CTkFont(size=11, weight="bold"), command=self.navigate_prev
        )
        self.btn_prev.pack(side="left", padx=(0, 6))

        self.lbl_page_num = ctk.CTkLabel(
            bottom_row, text="Page 1 of 1",
            font=ctk.CTkFont(size=12, weight="bold"), text_color=COL_TEXT_MUTED
        )
        self.lbl_page_num.pack(side="left", expand=True)

        self.btn_next = ctk.CTkButton(
            bottom_row, text="Next ➡", width=90, height=32,
            fg_color=COL_BG_CARD, border_color=COL_BORDER, border_width=1,
            text_color=COL_TEXT_MUTED, hover_color=COL_BG_CARD_ALT,
            font=ctk.CTkFont(size=11, weight="bold"), command=self.navigate_next
        )
        self.btn_next.pack(side="right", padx=(6, 0))

        # Export button — prominent green
        self.btn_export = ctk.CTkButton(
            bottom_row, text="📊  Export to Excel / CSV", height=32, width=170,
            corner_radius=6, fg_color=COL_SUCCESS_BG, hover_color="#1c3328",
            text_color=COL_SUCCESS, border_color="#2d5a3b", border_width=1,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.export_to_csv
        )
        self.btn_export.pack(side="right", padx=(6, 0))

        # Export as PDF button
        self.btn_export_pdf = ctk.CTkButton(
            bottom_row, text="📄  Export as PDF", height=32, width=150,
            corner_radius=6, fg_color=COL_DANGER_BG, hover_color="#3a1c1c",
            text_color=COL_DANGER, border_color="#5a2d2d", border_width=1,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.export_to_pdf
        )
        self.btn_export_pdf.pack(side="right", padx=(6, 0))

        # =================================================================
        # 📋 RIGHT PANEL: INVOICE DETAIL & ACTIONS
        # =================================================================
        self.right_pane = ctk.CTkFrame(
            self, fg_color=COL_BG_CARD,
            border_color=COL_BORDER, border_width=1, corner_radius=10
        )
        self.right_pane.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        ctk.CTkLabel(
            self.right_pane, text="Invoice Details",
            font=ctk.CTkFont(size=16, weight="bold"), text_color=COL_TEXT_MAIN
        ).pack(anchor="w", padx=22, pady=(22, 4))

        ctk.CTkLabel(
            self.right_pane, text="Select an invoice from the list to view details",
            font=ctk.CTkFont(size=11), text_color=COL_TEXT_SOFT
        ).pack(anchor="w", padx=22, pady=(0, 18))

        ctk.CTkFrame(self.right_pane, fg_color=COL_BORDER, height=1).pack(fill="x", padx=22)

        # Scrollable container for detail rows + action buttons, so they
        # remain accessible even on smaller window heights.
        self.right_pane_scroll = ctk.CTkScrollableFrame(self.right_pane, fg_color="transparent")
        self.right_pane_scroll.pack(fill="both", expand=True)

        self.card_inv_id    = self._make_detail_node("Invoice Number", "#INV-0000")
        self.card_cust      = self._make_detail_node("Customer", "—")
        self.card_date      = self._make_detail_node("Date & Time", "—")
        self.card_mode      = self._make_detail_node("Payment Mode", "—")
        self.card_subtotal  = self._make_detail_node("Subtotal", "Rs. 0.00")
        self.card_discount  = self._make_detail_node("Discount Applied", "Rs. 0.00", color="#f0ad4e")
        self.card_tax       = self._make_detail_node("Tax Applied", "0%")
        self.card_net       = self._make_detail_node("Net Payable", "Rs. 0.00", highlight=True)
        self.card_profit    = self._make_detail_node("Profit on Invoice", "Rs. —", highlight=False, color="#a855f7")

        ctk.CTkFrame(self.right_pane_scroll, fg_color=COL_BORDER, height=1).pack(fill="x", padx=22, pady=(10, 0))

        # Action buttons
        self.btn_print_action = ctk.CTkButton(
            self.right_pane_scroll, text="🖨️  Open Print Preview", height=44,
            corner_radius=8, fg_color=COL_DANGER_BG, hover_color="#3a1c1c",
            text_color=COL_DANGER, border_color="#5a2d2d", border_width=1,
            font=ctk.CTkFont(size=13, weight="bold"),
            state="disabled", command=self.execute_thermal_print
        )
        self.btn_print_action.pack(fill="x", padx=22, pady=(18, 8))

        self.btn_pdf_action = ctk.CTkButton(
            self.right_pane_scroll, text="📥  Export This Invoice PDF", height=40,
            corner_radius=8, fg_color="transparent",
            border_color=COL_ACCENT, border_width=1,
            text_color=COL_ACCENT, hover_color=COL_ACCENT_SOFT,
            font=ctk.CTkFont(size=12, weight="bold"),
            state="disabled", command=self.export_pdf_document
        )
        self.btn_pdf_action.pack(fill="x", padx=22, pady=(0, 8))

        self.btn_export_single = ctk.CTkButton(
            self.right_pane_scroll, text="📄  Copy Invoice Summary", height=36,
            corner_radius=8, fg_color="transparent",
            border_color=COL_BORDER, border_width=1,
            text_color=COL_TEXT_SOFT, hover_color=COL_BG_CARD_ALT,
            font=ctk.CTkFont(size=12),
            state="disabled", command=self.copy_invoice_summary
        )
        self.btn_export_single.pack(fill="x", padx=22, pady=(0, 22))

        # Load data
        self.fetch_all_records_to_cache()

    # ─────────────────────────────────────────────────────────────────
    # HELPER: STAT TILE
    # ─────────────────────────────────────────────────────────────────
    def _make_stat_tile(self, parent, label, value, color):
        tile = ctk.CTkFrame(parent, fg_color=COL_BG_CARD,
                             border_color=COL_BORDER, border_width=1, corner_radius=8)
        tile.pack(side="left", expand=True, fill="x", padx=3, ipady=6)
        ctk.CTkLabel(tile, text=label, font=ctk.CTkFont(size=10), text_color=COL_TEXT_SOFT).pack(pady=(6, 0))
        val_lbl = ctk.CTkLabel(tile, text=value, font=ctk.CTkFont(size=14, weight="bold"), text_color=color)
        val_lbl.pack(pady=(2, 6))
        return val_lbl

    # ─────────────────────────────────────────────────────────────────
    # HELPER: DETAIL NODE (RIGHT PANEL)
    # ─────────────────────────────────────────────────────────────────
    def _make_detail_node(self, title, default, highlight=False, color=None):
        f = ctk.CTkFrame(self.right_pane_scroll, fg_color="transparent")
        f.pack(fill="x", padx=22, pady=6)
        ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=11),
                     text_color=COL_TEXT_SOFT).pack(anchor="w")
        size  = 20 if highlight else 13
        if color:
            node_color = color
        else:
            node_color = COL_SUCCESS if highlight else COL_TEXT_MAIN
        lbl = ctk.CTkLabel(f, text=default,
                            font=ctk.CTkFont(size=size, weight="bold"),
                            text_color=node_color)
        lbl.pack(anchor="w", pady=(1, 0))
        return lbl

    # ─────────────────────────────────────────────────────────────────
    # FILTER MODE SWITCHING
    # ─────────────────────────────────────────────────────────────────
    def switch_filter_mode(self, mode):
        self.filter_mode = mode
        if mode == "date":
            self.btn_tab_date.configure(fg_color=COL_ACCENT, text_color="#fff",
                                         border_width=0)
            self.btn_tab_month.configure(fg_color=COL_BG_INPUT, text_color=COL_TEXT_MUTED,
                                          border_color=COL_BORDER, border_width=1)
            self.month_filter_frame.pack_forget()
            self.date_filter_frame.pack(fill="x", padx=16, pady=(0, 14))
        else:
            self.btn_tab_month.configure(fg_color=COL_ACCENT, text_color="#fff",
                                          border_width=0)
            self.btn_tab_date.configure(fg_color=COL_BG_INPUT, text_color=COL_TEXT_MUTED,
                                         border_color=COL_BORDER, border_width=1)
            self.date_filter_frame.pack_forget()
            self.month_filter_frame.pack(fill="x", padx=16, pady=(0, 14))
        self.trigger_new_search_context()

    # ─────────────────────────────────────────────────────────────────
    # DATA FETCHING
    # ─────────────────────────────────────────────────────────────────
    def fetch_all_records_to_cache(self):
        s_txt = self.search_entry.get().strip()

        if self.filter_mode == "date":
            start = self.inp_start_date.get_date().strftime('%Y-%m-%d')
            end   = self.inp_end_date.get_date().strftime('%Y-%m-%d')
        else:
            # Monthly mode — build start/end from month+year selection
            month_idx = MONTHS.index(self.month_var.get()) + 1
            year      = int(self.year_var.get())
            import calendar
            last_day  = calendar.monthrange(year, month_idx)[1]
            start     = f"{year}-{month_idx:02d}-01"
            end       = f"{year}-{month_idx:02d}-{last_day:02d}"

        self.cached_invoices = db.get_invoices_with_profit(s_txt, start, end)
        self._update_stats()
        self.render_paginated_table()

    def _update_stats(self):
        total = len(self.cached_invoices)
        revenue = sum(inv[4] for inv in self.cached_invoices) if self.cached_invoices else 0
        avg = revenue / total if total > 0 else 0
        # Profit: inv[7] is None for old invoices (no invoice_items saved)
        profits = [inv[7] for inv in self.cached_invoices if inv[7] is not None]
        total_profit = sum(profits) if profits else None
        self.stat_count.configure(text=str(total))
        self.stat_revenue.configure(text=f"Rs. {revenue:,.0f}")
        if total_profit is not None:
            self.stat_profit.configure(text=f"Rs. {total_profit:,.0f}")
        else:
            self.stat_profit.configure(text="Rs. —")
        self.stat_avg.configure(text=f"Rs. {avg:,.0f}")

    def trigger_new_search_context(self):
        self.current_page = 0
        self.fetch_all_records_to_cache()

    # ─────────────────────────────────────────────────────────────────
    # TABLE RENDERING
    # ─────────────────────────────────────────────────────────────────
    def render_paginated_table(self):
        for w in self.table_container.winfo_children():
            w.destroy()

        total_items = len(self.cached_invoices)
        max_pages = max(1, (total_items + self.items_per_page - 1) // self.items_per_page)

        self.lbl_page_num.configure(text=f"Page {self.current_page + 1} of {max_pages}")
        self.btn_prev.configure(state="normal" if self.current_page > 0 else "disabled")
        self.btn_next.configure(state="normal" if (self.current_page + 1) < max_pages else "disabled")

        start_idx = self.current_page * self.items_per_page
        page_chunk = self.cached_invoices[start_idx:start_idx + self.items_per_page]

        # Column widths
        W = [75, 160, 110, 90, 70, 55]

        # Header
        h = ctk.CTkFrame(self.table_container, fg_color=COL_BG_CARD_ALT,
                          corner_radius=6, height=34)
        h.pack(fill="x", padx=6, pady=(6, 4))

        for txt, w, anchor in [
            ("Invoice",   W[0], "center"),
            ("Customer",  W[1], "w"),
            ("Date",      W[2], "center"),
            ("Mode",      W[3], "center"),
            ("Amount",    W[4], "center"),
            ("",          W[5], "center"),
        ]:
            ctk.CTkLabel(h, text=txt, width=w, font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=COL_TEXT_SOFT, anchor=anchor).pack(side="left", padx=4, pady=6)

        if not page_chunk:
            ctk.CTkLabel(self.table_container, text="No invoices found for the selected filter.",
                         font=ctk.CTkFont(size=12), text_color=COL_TEXT_SOFT).pack(pady=30)
            return

        for idx, inv in enumerate(page_chunk):
            inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = inv

            row_bg = COL_BG_CARD_ALT if idx % 2 == 0 else "transparent"
            row = ctk.CTkFrame(self.table_container, fg_color=row_bg, corner_radius=6, height=40)
            row.pack(fill="x", padx=6, pady=2)

            ctk.CTkLabel(row, text=f"#{inv_id:04d}", width=W[0],
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=COL_ACCENT, anchor="center").pack(side="left", padx=4, pady=6)

            name_disp = (cust_name or "Walk-in")
            if len(name_disp) > 20: name_disp = name_disp[:18] + "…"
            ctk.CTkLabel(row, text=name_disp, width=W[1],
                         font=ctk.CTkFont(size=12), text_color=COL_TEXT_MAIN,
                         anchor="w").pack(side="left", padx=4)

            date_disp = timestamp[:10] if timestamp else "—"
            ctk.CTkLabel(row, text=date_disp, width=W[2],
                         font=ctk.CTkFont(size=11), text_color=COL_TEXT_MUTED,
                         anchor="center").pack(side="left", padx=4)

            mode_color = COL_SUCCESS if pay_mode == "CASH" else COL_WARNING
            mode_text  = "💵 Cash" if pay_mode == "CASH" else "💳 Khata"
            ctk.CTkLabel(row, text=mode_text, width=W[3],
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=mode_color, anchor="center").pack(side="left", padx=4)

            ctk.CTkLabel(row, text=f"Rs.{net_amount:,.0f}", width=W[4],
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=COL_TEXT_MAIN, anchor="center").pack(side="left", padx=4)

            ctk.CTkButton(
                row, text="View", width=W[5], height=26, corner_radius=4,
                fg_color=COL_ACCENT_SOFT, hover_color=COL_BORDER,
                text_color=COL_ACCENT, font=ctk.CTkFont(size=11, weight="bold"),
                command=lambda p=inv: self.focus_invoice_target(p)
            ).pack(side="right", padx=6)

    # ─────────────────────────────────────────────────────────────────
    # PAGINATION
    # ─────────────────────────────────────────────────────────────────
    def navigate_prev(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.render_paginated_table()

    def navigate_next(self):
        max_pages = (len(self.cached_invoices) + self.items_per_page - 1) // self.items_per_page
        if (self.current_page + 1) < max_pages:
            self.current_page += 1
            self.render_paginated_table()

    def clear_date_inputs(self):
        today = datetime.date.today()
        self.inp_start_date.set_date(today.replace(day=1))
        self.inp_end_date.set_date(today)
        self.trigger_new_search_context()

    # ─────────────────────────────────────────────────────────────────
    # INVOICE DETAIL (RIGHT PANEL)
    # ─────────────────────────────────────────────────────────────────
    def focus_invoice_target(self, payload):
        self.active_selection = payload
        inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = payload

        self.card_inv_id.configure(text=f"#INV-{inv_id:04d}")
        self.card_cust.configure(text=cust_name if cust_name else "Walk-in Customer")
        self.card_date.configure(text=timestamp or "—")
        mode_text = "💵 Cash" if pay_mode == "CASH" else "💳 Khata / Wallet"
        self.card_mode.configure(text=mode_text)
        self.card_subtotal.configure(text=f"Rs. {subtotal:,.2f}")
        if discount_amount and float(discount_amount) > 0:
            note = discount_note or "Discount"
            self.card_discount.configure(text=f"- Rs. {float(discount_amount):,.2f}  ({note})")
        else:
            self.card_discount.configure(text="Rs. 0.00")
        self.card_tax.configure(text=f"{tax_pct:g}%  →  Rs. {(subtotal * float(tax_pct) / 100):,.2f}")
        self.card_net.configure(text=f"Rs. {net_amount:,.2f}")

        # Profit per invoice
        if profit is not None:
            self.card_profit.configure(text=f"Rs. {profit:,.2f}")
        else:
            # Old invoice — calculate from items if available
            items = db.get_invoice_profit(inv_id)
            if items:
                total_p = sum(row[5] for row in items)
                self.card_profit.configure(text=f"Rs. {total_p:,.2f}")
            else:
                self.card_profit.configure(text="Rs. — (purana invoice)")

        for btn in [self.btn_print_action, self.btn_pdf_action, self.btn_export_single]:
            btn.configure(state="normal")

    # ─────────────────────────────────────────────────────────────────
    # ACTIONS
    # ─────────────────────────────────────────────────────────────────
    def execute_thermal_print(self):
        if not self.active_selection: return
        inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = self.active_selection
        items = db.get_invoice_items_for_print(inv_id)
        if not items:
            items = [[0, "Transaction Settlement Record", 1, subtotal, subtotal, 1]]
        InvoicePrintWindow(self, inv_id, cust_name or "Walk-in Customer",
                           timestamp, subtotal, tax_pct, net_amount, items,
                           discount_amount=discount_amount, discount_note=discount_note)

    def export_pdf_document(self):
        if not self.active_selection: return
        inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = self.active_selection
        items = db.get_invoice_items_for_print(inv_id)
        if not items:
            items = [[0, "Transaction Settlement Record", 1, subtotal, subtotal, 1]]

        default_name = f"invoice_INV-{inv_id:04d}.pdf"
        save_path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Invoice PDF As",
            initialfile=default_name,
            defaultextension=".pdf",
            filetypes=(("PDF Document", "*.pdf"), ("All Files", "*.*"))
        )
        if not save_path:
            return

        success, msg = export_invoice_pdf(
            save_path, inv_id, cust_name or "Walk-in Customer",
            timestamp, subtotal, tax_pct, net_amount, items,
            discount_amount=discount_amount, discount_note=discount_note
        )
        if success:
            messagebox.showinfo("PDF Export", msg)
        else:
            messagebox.showerror("PDF Export Failed", msg)

    def copy_invoice_summary(self):
        if not self.active_selection: return
        inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = self.active_selection
        profit_str = f"Rs. {profit:,.2f}" if profit is not None else "—"
        discount_line = f"Discount: -Rs. {float(discount_amount):,.2f} ({discount_note})\n" if discount_amount and float(discount_amount) > 0 else ""
        summary = (
            f"Invoice: #INV-{inv_id:04d}\n"
            f"Customer: {cust_name or 'Walk-in Customer'}\n"
            f"Date: {timestamp}\n"
            f"Mode: {pay_mode}\n"
            f"Subtotal: Rs. {subtotal:,.2f}\n"
            f"{discount_line}"
            f"Tax: {tax_pct}%\n"
            f"Net Payable: Rs. {net_amount:,.2f}\n"
            f"Profit: {profit_str}"
        )
        self.clipboard_clear()
        self.clipboard_append(summary)
        messagebox.showinfo("Copied!", "Invoice summary copied to clipboard.")

    # ─────────────────────────────────────────────────────────────────
    # 📊 EXPORT TO EXCEL (with Profit)
    # ─────────────────────────────────────────────────────────────────
    def export_to_csv(self):
        if not self.cached_invoices:
            messagebox.showwarning("No Data", "No invoices to export. Apply a filter first.")
            return

        # Build suggested filename
        if self.filter_mode == "month":
            month_name = self.month_var.get()
            year = self.year_var.get()
            default_name = f"Invoices_{month_name}_{year}"
        else:
            start = self.inp_start_date.get_date().strftime('%Y%m%d')
            end   = self.inp_end_date.get_date().strftime('%Y%m%d')
            default_name = f"Invoices_{start}_to_{end}"

        if XLSX_AVAILABLE:
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx"), ("CSV File", "*.csv"), ("All Files", "*.*")],
                initialfile=default_name + ".xlsx",
                title="Export Invoices — Save As"
            )
        else:
            save_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
                initialfile=default_name + ".csv",
                title="Export Invoices — Save As"
            )

        if not save_path:
            return

        try:
            if save_path.endswith(".xlsx") and XLSX_AVAILABLE:
                self._export_xlsx(save_path)
            else:
                self._export_csv_fallback(save_path)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save file:\n{str(e)}")

    # ─────────────────────────────────────────────────────────────────
    # 📄 EXPORT AS PDF (Invoice Summary, with Profit)
    # ─────────────────────────────────────────────────────────────────
    def export_to_pdf(self):
        if not self.cached_invoices:
            messagebox.showwarning("No Data", "No invoices to export. Apply a filter first.")
            return

        # Build suggested filename + range label (same logic as Excel export)
        if self.filter_mode == "month":
            month_name = self.month_var.get()
            year = self.year_var.get()
            default_name = f"Invoices_{month_name}_{year}"
            range_label = f"{month_name} {year}"
        else:
            start_dt = self.inp_start_date.get_date()
            end_dt = self.inp_end_date.get_date()
            default_name = f"Invoices_{start_dt.strftime('%Y%m%d')}_to_{end_dt.strftime('%Y%m%d')}"
            range_label = f"{start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')}"

        save_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF File", "*.pdf"), ("All Files", "*.*")],
            initialfile=default_name + ".pdf",
            title="Export Invoices as PDF — Save As"
        )
        if not save_path:
            return

        try:
            headers = ["Invoice No", "Customer", "Date & Time", "Payment Mode",
                       "Subtotal", "Tax %", "Net Amount", "Profit", "Margin %", "Discount"]
            col_classes = ["center", "", "center", "center", "right", "center", "right", "right", "right", "right"]
            col_widths = [9, 17, 14, 11, 10, 6, 10, 9, 7, 7]

            rows = []
            total_revenue = 0
            total_profit_sum = 0
            profit_count = 0

            for inv in self.cached_invoices:
                inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = inv
                total_revenue += net_amount

                if profit is not None:
                    total_profit_sum += profit
                    profit_count += 1
                    margin = (profit / subtotal * 100) if subtotal else 0
                    profit_cell = (f"{profit:,.2f}", "blue")
                    margin_cell = f"{margin:.1f}%"
                else:
                    profit_cell = "—"
                    margin_cell = "—"

                rows.append([
                    f"INV-{inv_id:04d}",
                    cust_name or "Walk-in Customer",
                    timestamp,
                    pay_mode,
                    f"{subtotal:,.2f}",
                    f"{tax_pct}%",
                    f"{net_amount:,.2f}",
                    profit_cell,
                    margin_cell,
                    f"{float(discount_amount):,.2f}" if discount_amount else "0.00",
                ])

            totals_row = [
                ("TOTALS", "bold"), "", "", "", "", "",
                (f"{total_revenue:,.2f}", "green"),
                (f"{total_profit_sum:,.2f}", "blue") if profit_count > 0 else "—",
                "", "",
            ]

            avg_invoice = total_revenue / len(self.cached_invoices) if self.cached_invoices else 0

            summary_rows = [
                ("Report Range:", range_label),
                ("Total Invoices:", str(len(self.cached_invoices))),
                ("Total Revenue:", f"Rs. {total_revenue:,.2f}"),
                ("Total Profit:", f"Rs. {total_profit_sum:,.2f}" if profit_count > 0 else "—"),
                ("Average Invoice:", f"Rs. {avg_invoice:,.2f}"),
            ]

            html = build_table_html(
                title="📊 Invoice & Profit Report",
                subtitle=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                headers=headers,
                rows=rows,
                col_classes=col_classes,
                orientation="landscape",
                summary_rows=summary_rows,
                totals_row=totals_row,
                col_widths=col_widths,
            )

            success, msg = export_html_to_pdf(html, save_path)
            if success:
                messagebox.showinfo(
                    "Export Successful ✅",
                    f"{len(self.cached_invoices)} invoices exported to PDF!\n\nFile saved at:\n{save_path}"
                )
                try:
                    os.startfile(save_path)
                except Exception:
                    pass
            else:
                messagebox.showerror("Export Failed", msg)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save PDF:\n{str(e)}")

    def _export_xlsx(self, save_path):
        """Full professional Excel export with profit columns and per-invoice item breakdown"""
        wb = openpyxl.Workbook()

        # ── SHEET 1: Invoice Summary ──────────────────────────────
        ws1 = wb.active
        ws1.title = "Invoice Summary"

        # Styles
        hdr_fill   = PatternFill("solid", fgColor="1e3a5f")
        hdr_font   = Font(color="FFFFFF", bold=True, size=11, name="Arial")
        title_font = Font(color="FFFFFF", bold=True, size=14, name="Arial")
        profit_fill = PatternFill("solid", fgColor="2d1b4e")
        profit_font = Font(color="a855f7", bold=True, name="Arial")
        total_fill  = PatternFill("solid", fgColor="14241c")
        total_font  = Font(color="22c55e", bold=True, name="Arial")
        center_aln  = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="2a2e38"),
            right=Side(style="thin", color="2a2e38"),
            top=Side(style="thin", color="2a2e38"),
            bottom=Side(style="thin", color="2a2e38")
        )

        # Title row
        ws1.merge_cells("A1:J1")
        title_cell = ws1["A1"]
        title_cell.value = "📊 Invoice & Profit Report"
        title_cell.font = title_font
        title_cell.fill = PatternFill("solid", fgColor="0f1117")
        title_cell.alignment = center_aln
        ws1.row_dimensions[1].height = 30

        # Generated info
        ws1.merge_cells("A2:J2")
        info_cell = ws1["A2"]
        info_cell.value = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Total Invoices: {len(self.cached_invoices)}"
        info_cell.font = Font(color="94a3b8", italic=True, name="Arial", size=10)
        info_cell.fill = PatternFill("solid", fgColor="11131a")
        info_cell.alignment = center_aln
        ws1.row_dimensions[2].height = 20

        # Header row
        headers = ["Invoice No", "Customer Name", "Date & Time", "Payment Mode",
                   "Subtotal (Rs.)", "Tax %", "Net Amount (Rs.)", "Profit (Rs.)", "Margin %", "Discount (Rs.)"]
        ws1.row_dimensions[3].height = 22
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col_idx, value=hdr)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center_aln
            cell.border = thin_border

        # Data rows
        total_revenue = 0
        total_profit_sum = 0
        profit_count = 0

        for row_idx, inv in enumerate(self.cached_invoices, 4):
            inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = inv
            total_revenue += net_amount

            margin = None
            if profit is not None:
                total_profit_sum += profit
                profit_count += 1
                margin = (profit / subtotal * 100) if subtotal else 0

            row_fill = PatternFill("solid", fgColor="1a1d24") if row_idx % 2 == 0 else PatternFill("solid", fgColor="15171d")

            row_data = [
                f"INV-{inv_id:04d}",
                cust_name or "Walk-in Customer",
                timestamp,
                pay_mode,
                round(subtotal, 2),
                f"{tax_pct}%",
                round(net_amount, 2),
                round(profit, 2) if profit is not None else "—",
                f"{margin:.1f}%" if margin is not None else "—",
                round(float(discount_amount), 2) if discount_amount else 0
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.font = Font(name="Arial", size=10, color="f1f5f9")
                cell.border = thin_border
                if col_idx in (5, 7, 10):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 8 and profit is not None:
                    cell.font = profit_font
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in (1, 4, 6, 9):
                    cell.alignment = center_aln

        # Summary footer
        last_data_row = 3 + len(self.cached_invoices)
        footer_row = last_data_row + 2

        ws1.row_dimensions[footer_row].height = 22
        ws1.merge_cells(f"A{footer_row}:D{footer_row}")
        ws1.cell(row=footer_row, column=1, value="TOTALS").font = Font(bold=True, color="94a3b8", name="Arial")
        ws1.cell(row=footer_row, column=1).fill = PatternFill("solid", fgColor="11131a")
        ws1.cell(row=footer_row, column=1).alignment = center_aln

        rev_cell = ws1.cell(row=footer_row, column=7, value=round(total_revenue, 2))
        rev_cell.font = total_font
        rev_cell.fill = total_fill
        rev_cell.number_format = '#,##0.00'
        rev_cell.alignment = Alignment(horizontal="right")

        profit_total_cell = ws1.cell(row=footer_row, column=8,
                                      value=round(total_profit_sum, 2) if profit_count > 0 else "—")
        profit_total_cell.font = Font(bold=True, color="a855f7", name="Arial")
        profit_total_cell.fill = PatternFill("solid", fgColor="1a0a2e")
        if profit_count > 0:
            profit_total_cell.number_format = '#,##0.00'
        profit_total_cell.alignment = Alignment(horizontal="right")

        if profit_count > 0 and total_revenue > 0:
            overall_margin = total_profit_sum / total_revenue * 100
            ws1.cell(row=footer_row, column=9, value=f"{overall_margin:.1f}%").font = Font(bold=True, color="a855f7", name="Arial")

        # Column widths
        col_widths = [14, 25, 20, 14, 16, 8, 18, 16, 12, 16]
        for i, w in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        # ── SHEET 2: Per-Invoice Item Breakdown ──────────────────
        ws2 = wb.create_sheet("Item Breakdown")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:H1")
        ws2["A1"].value = "📦 Per-Invoice Item Breakdown with Profit"
        ws2["A1"].font = title_font
        ws2["A1"].fill = PatternFill("solid", fgColor="0f1117")
        ws2["A1"].alignment = center_aln
        ws2.row_dimensions[1].height = 28

        item_headers = ["Invoice No", "Customer", "Product", "Qty", "Cost Price", "Sale Price", "Line Total", "Item Profit"]
        ws2.row_dimensions[2].height = 22
        for ci, h in enumerate(item_headers, 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center_aln
            c.border = thin_border

        current_row = 3
        for inv in self.cached_invoices:
            inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = inv
            items = db.get_invoice_profit(inv_id)
            if not items:
                # Old invoice — no items saved
                r_fill = PatternFill("solid", fgColor="1a1d24") if current_row % 2 == 0 else PatternFill("solid", fgColor="15171d")
                for ci, val in enumerate([f"INV-{inv_id:04d}", cust_name or "Walk-in", "—", "—", "—", "—", "—", "—"], 1):
                    c = ws2.cell(row=current_row, column=ci, value=val)
                    c.fill = r_fill
                    c.font = Font(name="Arial", size=10, color="64748b", italic=True)
                    c.border = thin_border
                current_row += 1
            else:
                for item in items:
                    prod_name, qty, cost_p, sale_p, line_total, item_profit = item
                    r_fill = PatternFill("solid", fgColor="1a1d24") if current_row % 2 == 0 else PatternFill("solid", fgColor="15171d")
                    row_data = [
                        f"INV-{inv_id:04d}",
                        cust_name or "Walk-in",
                        prod_name,
                        qty,
                        round(cost_p, 2),
                        round(sale_p, 2),
                        round(line_total, 2),
                        round(item_profit, 2)
                    ]
                    for ci, val in enumerate(row_data, 1):
                        c = ws2.cell(row=current_row, column=ci, value=val)
                        c.fill = r_fill
                        c.font = Font(name="Arial", size=10, color="f1f5f9")
                        c.border = thin_border
                        if ci in (5, 6, 7):
                            c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal="right")
                        elif ci == 8:
                            c.font = Font(name="Arial", size=10, color="a855f7", bold=True)
                            c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal="right")
                        elif ci == 4:
                            c.alignment = center_aln
                    current_row += 1

        item_col_widths = [14, 22, 28, 8, 14, 14, 14, 14]
        for i, w in enumerate(item_col_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        wb.save(save_path)
        messagebox.showinfo(
            "Export Successful ✅",
            f"{len(self.cached_invoices)} invoices exported successfully!\n\n"
            f"• Sheet 1: Invoice Summary (with profit)\n"
            f"• Sheet 2: Per-item breakdown\n\n"
            f"File saved at:\n{save_path}"
        )
        try:
            os.startfile(save_path)
        except Exception:
            pass

    def _export_csv_fallback(self, save_path):
        """CSV fallback if openpyxl not installed"""
        with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Invoice No", "Customer Name", "Subtotal (Rs.)",
                "Tax %", "Net Amount (Rs.)", "Date & Time", "Payment Mode", "Profit (Rs.)", "Discount (Rs.)"
            ])
            for inv in self.cached_invoices:
                inv_id, cust_name, subtotal, tax_pct, net_amount, timestamp, pay_mode, profit, discount_amount, discount_note = inv
                writer.writerow([
                    f"INV-{inv_id:04d}",
                    cust_name or "Walk-in Customer",
                    f"{subtotal:.2f}",
                    f"{tax_pct}%",
                    f"{net_amount:.2f}",
                    timestamp,
                    pay_mode,
                    f"{profit:.2f}" if profit is not None else "—",
                    f"{float(discount_amount):.2f}" if discount_amount else "0.00"
                ])
            writer.writerow([])
            total_revenue = sum(inv[4] for inv in self.cached_invoices)
            profits = [inv[7] for inv in self.cached_invoices if inv[7] is not None]
            writer.writerow(["", "TOTAL INVOICES", len(self.cached_invoices), "", "", "", "", "", ""])
            writer.writerow(["", "TOTAL REVENUE", "", "", f"{total_revenue:.2f}", "", "", "", ""])
            if profits:
                writer.writerow(["", "TOTAL PROFIT", "", "", "", "", "", f"{sum(profits):.2f}", ""])

        messagebox.showinfo(
            "Export Successful ✅",
            f"{len(self.cached_invoices)} invoices exported!\n\nFile saved at:\n{save_path}"
        )
        try:
            os.startfile(save_path)
        except Exception:
            pass
